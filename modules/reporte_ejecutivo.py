"""
Reporte Ejecutivo
Consolida todos los módulos en un único Excel para presentar a gerencia.
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _borde_medio():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center", wrap_text=True)
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")

VERDE      = "E2EFDA"
AMARILLO   = "FFF2CC"
ROJO       = "FFCCCC"
AZUL_OSC   = "1F3864"
AZUL_MED   = "2E75B6"
AZUL_CLAR  = "D6E4F0"
GRIS       = "F2F2F2"
VERDE_T    = "375623"
AMARILLO_T = "7F6000"
ROJO_T     = "C00000"


# ── Helpers ────────────────────────────────────────────────────────

def celda(ws, ref, valor, fuente=None, relleno=None, alin=None, borde=None, fmt=None):
    c = ws[ref] if isinstance(ref, str) else ws.cell(ref[0], ref[1])
    c.value = valor
    if fuente:  c.font           = fuente
    if relleno: c.fill           = _fill(relleno)
    if alin:    c.alignment      = alin
    if borde:   c.border         = borde
    if fmt:     c.number_format  = fmt
    return c


def kpi_box(ws, fila, col, titulo, valor, unidad, semaforo):
    color   = VERDE if semaforo == "verde" else AMARILLO if semaforo == "amarillo" else ROJO
    color_t = VERDE_T if semaforo == "verde" else AMARILLO_T if semaforo == "amarillo" else ROJO_T
    icono   = "✅" if semaforo == "verde" else "⚠️" if semaforo == "amarillo" else "🔴"

    ws.cell(fila, col, titulo).font      = Font(name="Arial", bold=True, size=9, color="555555")
    ws.cell(fila, col).fill              = _fill(GRIS)
    ws.cell(fila, col).alignment         = CENTRO
    ws.cell(fila, col).border            = _borde()

    ws.cell(fila+1, col, f"{icono} {valor} {unidad}").font      = Font(name="Arial", bold=True, size=11, color=color_t)
    ws.cell(fila+1, col).fill            = _fill(color)
    ws.cell(fila+1, col).alignment       = CENTRO
    ws.cell(fila+1, col).border          = _borde()


# ── Carga de datos ─────────────────────────────────────────────────

def cargar_cobranzas(ruta):
    try:
        df = pd.read_excel(ruta)
        df.columns = df.columns.str.strip().str.lower()
        df["fecha_venc"]   = pd.to_datetime(df["fecha_venc"], dayfirst=True)
        hoy                = pd.Timestamp(date.today())
        df["dias_vencido"] = (hoy - df["fecha_venc"]).dt.days
        def tramo(d):
            if d <= 0:    return "Al día"
            elif d <= 30: return "0-30"
            elif d <= 60: return "31-60"
            elif d <= 90: return "61-90"
            else:         return "+90"
        df["tramo"] = df["dias_vencido"].apply(tramo)
        PROB = {"Al día": 0.95, "0-30": 0.85, "31-60": 0.65, "61-90": 0.40, "+90": 0.15}
        df["monto_estimado"] = df["monto"] * df["tramo"].map(PROB)
        return df
    except:
        return None


def cargar_proveedores(ruta):
    try:
        df = pd.read_excel(ruta)
        df.columns = df.columns.str.strip().str.lower()
        df["fecha_factura"] = pd.to_datetime(df["fecha_factura"], dayfirst=True)
        df["fecha_pago"]    = pd.to_datetime(df["fecha_pago"], dayfirst=True, errors="coerce")
        df["dias_pago"]     = (df["fecha_pago"] - df["fecha_factura"]).dt.days
        df["credito_dias"]  = df.get("credito_dias", 30).fillna(30)
        return df
    except:
        return None


def cargar_lineas(ruta):
    try:
        df = pd.read_excel(ruta)
        df.columns = df.columns.str.strip().str.lower()
        df["fecha_vencimiento"] = pd.to_datetime(df["fecha_vencimiento"], dayfirst=True, errors="coerce")
        df["cupo_total"]        = pd.to_numeric(df["cupo_total"],  errors="coerce")
        df["cupo_usado"]        = pd.to_numeric(df["cupo_usado"],  errors="coerce")
        df["cupo_disponible"]   = df["cupo_total"] - df["cupo_usado"]
        df["pct_uso"]           = df["cupo_usado"] / df["cupo_total"]
        hoy = pd.Timestamp(date.today())
        df["dias_vencimiento"]  = (df["fecha_vencimiento"] - hoy).dt.days
        return df
    except:
        return None


def cargar_indicadores(ruta):
    try:
        df_b = pd.read_excel(ruta, sheet_name="Balance")
        df_r = pd.read_excel(ruta, sheet_name="Resultados")
        df_b.columns = df_b.columns.str.strip().str.lower()
        df_r.columns = df_r.columns.str.strip().str.lower()
        df_b["cuenta"] = df_b["cuenta"].str.strip().str.lower()
        df_r["cuenta"] = df_r["cuenta"].str.strip().str.lower()
        b = df_b.set_index("cuenta")["valor_actual"].to_dict()
        r = df_r.set_index("cuenta")["valor_actual"].to_dict()
        return b, r
    except:
        return None, None


# ── Generación Excel ───────────────────────────────────────────────

def generar_reporte(empresa, periodo, rutas, archivo):
    wb = Workbook()

    # ── Hoja 1: Portada ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Portada"

    for i in range(1, 8):
        ws.row_dimensions[i].height = 40
    for i in range(1, 4):
        ws.column_dimensions[get_column_letter(i)].width = 30

    ws.merge_cells("A1:C1")
    ws["A1"] = empresa.upper()
    ws["A1"].font      = Font(name="Arial", bold=True, size=22, color="FFFFFF")
    ws["A1"].fill      = _fill(AZUL_OSC)
    ws["A1"].alignment = CENTRO

    ws.merge_cells("A2:C2")
    ws["A2"] = "REPORTE EJECUTIVO FINANCIERO"
    ws["A2"].font      = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws["A2"].fill      = _fill(AZUL_MED)
    ws["A2"].alignment = CENTRO

    ws.merge_cells("A3:C3")
    ws["A3"] = f"Período: {periodo}   |   Generado: {date.today().strftime('%d/%m/%Y')}"
    ws["A3"].font      = Font(name="Arial", size=11, color="555555")
    ws["A3"].fill      = _fill(AZUL_CLAR)
    ws["A3"].alignment = CENTRO

    # Índice
    ws.merge_cells("A5:C5")
    ws["A5"] = "CONTENIDO"
    ws["A5"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws["A5"].fill      = _fill(AZUL_OSC)
    ws["A5"].alignment = CENTRO

    contenido = [
        ("1", "Dashboard KPIs",          "Vista rápida del estado financiero"),
        ("2", "Cobranzas",               "Cartera, tramos y estimación de recaudación"),
        ("3", "Proveedores",             "Ranking, concentración y días de pago"),
        ("4", "Líneas de Crédito",       "Cupos, uso y vencimientos"),
        ("5", "Indicadores Financieros", "Ratios clave con semáforo de alertas"),
    ]

    for i, (num, titulo, desc) in enumerate(contenido, 6):
        ws.row_dimensions[i].height = 22
        ws.cell(i, 1, num).font       = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.cell(i, 1).fill            = _fill(AZUL_MED)
        ws.cell(i, 1).alignment       = CENTRO
        ws.cell(i, 1).border          = _borde()
        ws.cell(i, 2, titulo).font    = Font(name="Arial", bold=True, size=10)
        ws.cell(i, 2).fill            = _fill(AZUL_CLAR)
        ws.cell(i, 2).alignment       = IZQUIERDA
        ws.cell(i, 2).border          = _borde()
        ws.cell(i, 3, desc).font      = Font(name="Arial", size=10, color="555555")
        ws.cell(i, 3).fill            = _fill(GRIS)
        ws.cell(i, 3).alignment       = IZQUIERDA
        ws.cell(i, 3).border          = _borde()

    # ── Hoja 2: Dashboard KPIs ────────────────────────────────────
    ws2 = wb.create_sheet("Dashboard")
    for i in range(1, 20):
        ws2.row_dimensions[i].height = 24
    for i in range(1, 7):
        ws2.column_dimensions[get_column_letter(i)].width = 22

    ws2.merge_cells("A1:F1")
    ws2["A1"] = f"DASHBOARD EJECUTIVO — {periodo}"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws2["A1"].fill      = _fill(AZUL_OSC)
    ws2["A1"].alignment = CENTRO
    ws2.row_dimensions[1].height = 32

    # Sección cobranzas
    df_cob = cargar_cobranzas(rutas.get("cobranzas", ""))
    if df_cob is not None:
        total_cartera  = df_cob["monto"].sum()
        total_estimado = df_cob["monto_estimado"].sum()
        pct_rec        = total_estimado / total_cartera if total_cartera else 0
        criticos       = df_cob[df_cob["tramo"] == "+90"]["monto"].sum()
        sem_cob        = "verde" if pct_rec >= 0.7 else "amarillo" if pct_rec >= 0.5 else "rojo"

        ws2.merge_cells("A2:F2")
        ws2["A2"] = "💰 COBRANZAS"
        ws2["A2"].font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws2["A2"].fill      = _fill(AZUL_MED)
        ws2["A2"].alignment = IZQUIERDA

        kpi_box(ws2, 3, 1, "Cartera Total",         f"${total_cartera/1000000:.1f}M",  "",    "verde")
        kpi_box(ws2, 3, 2, "Estimado Recaudar",      f"${total_estimado/1000000:.1f}M", "",    sem_cob)
        kpi_box(ws2, 3, 3, "% Recuperación",         f"{pct_rec:.1%}",                  "",    sem_cob)
        kpi_box(ws2, 3, 4, "Cartera +90 días",       f"${criticos/1000000:.1f}M",       "",    "rojo" if criticos > 0 else "verde")
        kpi_box(ws2, 3, 5, "Clientes",               str(df_cob["cliente"].nunique()),   "",    "verde")
        kpi_box(ws2, 3, 6, "Facturas",               str(len(df_cob)),                   "",    "verde")

    # Sección líneas
    df_lin = cargar_lineas(rutas.get("lineas", ""))
    if df_lin is not None:
        cupo_total = df_lin["cupo_total"].sum()
        cupo_usado = df_lin["cupo_usado"].sum()
        cupo_disp  = df_lin["cupo_disponible"].sum()
        pct_uso    = cupo_usado / cupo_total if cupo_total else 0
        sem_lin    = "verde" if pct_uso < 0.5 else "amarillo" if pct_uso < 0.8 else "rojo"
        venc_crit  = len(df_lin[df_lin["dias_vencimiento"].fillna(999) < 30])

        ws2.merge_cells("A6:F6")
        ws2["A6"] = "🏦 LÍNEAS DE CRÉDITO"
        ws2["A6"].font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws2["A6"].fill      = _fill(AZUL_MED)
        ws2["A6"].alignment = IZQUIERDA

        kpi_box(ws2, 7, 1, "Cupo Total",            f"${cupo_total/1000000:.1f}M",  "", "verde")
        kpi_box(ws2, 7, 2, "Cupo Usado",            f"${cupo_usado/1000000:.1f}M",  "", sem_lin)
        kpi_box(ws2, 7, 3, "Cupo Disponible",       f"${cupo_disp/1000000:.1f}M",   "", sem_lin)
        kpi_box(ws2, 7, 4, "% Uso Global",          f"{pct_uso:.1%}",               "", sem_lin)
        kpi_box(ws2, 7, 5, "Líneas",                str(len(df_lin)),                "", "verde")
        kpi_box(ws2, 7, 6, "Vencen < 30 días",     str(venc_crit),                  "", "rojo" if venc_crit > 0 else "verde")

    # Sección indicadores
    b, r = cargar_indicadores(rutas.get("indicadores", ""))
    if b and r:
        def get(d, k): return d.get(k, 0) or 0
        ac = get(b, "activo_corriente")
        pc = get(b, "pasivo_corriente")
        liq = ac / pc if pc else 0
        ebitda = get(r, "ebitda")
        df_fin = get(b, "deuda_financiera")
        deu_ebitda = df_fin / ebitda if ebitda else 0
        margen = ebitda / get(r, "ingresos") if get(r, "ingresos") else 0

        ws2.merge_cells("A10:F10")
        ws2["A10"] = "📊 INDICADORES FINANCIEROS"
        ws2["A10"].font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws2["A10"].fill      = _fill(AZUL_MED)
        ws2["A10"].alignment = IZQUIERDA

        kpi_box(ws2, 11, 1, "Liquidez Corriente",   f"{liq:.2f}x",      "", "verde" if liq >= 1.5 else "amarillo" if liq >= 1.0 else "rojo")
        kpi_box(ws2, 11, 2, "Deuda / EBITDA",       f"{deu_ebitda:.2f}x","", "verde" if deu_ebitda <= 3 else "amarillo" if deu_ebitda <= 4.5 else "rojo")
        kpi_box(ws2, 11, 3, "Margen EBITDA",        f"{margen:.1%}",     "", "verde" if margen >= 0.15 else "amarillo" if margen >= 0.08 else "rojo")
        kpi_box(ws2, 11, 4, "Ingresos",             f"${get(r,'ingresos')/1000000:.1f}M", "", "verde")
        kpi_box(ws2, 11, 5, "EBITDA",               f"${ebitda/1000000:.1f}M",            "", "verde" if margen >= 0.15 else "amarillo")
        kpi_box(ws2, 11, 6, "Utilidad Neta",        f"${get(r,'utilidad_neta')/1000000:.1f}M", "", "verde" if get(r,'utilidad_neta') > 0 else "rojo")

    # ── Hoja 3: Cobranzas ─────────────────────────────────────────
    ws3 = wb.create_sheet("Cobranzas")
    if df_cob is not None:
        for i, w in enumerate([16, 16, 14, 16, 16], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        ws3.merge_cells("A1:E1")
        ws3["A1"] = "RESUMEN DE COBRANZAS"
        ws3["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws3["A1"].fill      = _fill(AZUL_OSC)
        ws3["A1"].alignment = CENTRO
        ws3.row_dimensions[1].height = 28

        for col, h in enumerate(["Tramo", "Facturas", "Monto Cartera", "Monto Estimado", "% Recuperación"], 1):
            c = ws3.cell(2, col, h)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", bold=True, color="FFFFFF", size=10),
                _fill(AZUL_OSC), CENTRO, _borde())

        PROB   = {"Al día": 0.95, "0-30": 0.85, "31-60": 0.65, "61-90": 0.40, "+90": 0.15}
        COLORS = {"Al día": VERDE, "0-30": "FFFFFF", "31-60": AMARILLO, "61-90": "FCE4D6", "+90": ROJO}
        tramos = ["Al día", "0-30", "31-60", "61-90", "+90"]
        total  = df_cob["monto"].sum()

        for i, tramo in enumerate(tramos, 3):
            sub    = df_cob[df_cob["tramo"] == tramo]
            monto  = sub["monto"].sum()
            est    = sub["monto_estimado"].sum()
            pct    = est / monto if monto else 0
            color  = COLORS.get(tramo, "FFFFFF")
            datos  = [tramo, len(sub), monto, est, pct]
            fmts   = [None, "#,##0", "#,##0", "#,##0", "0.0%"]
            alns   = [CENTRO, CENTRO, DERECHA, DERECHA, CENTRO]
            for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
                c = ws3.cell(i, col, v)
                c.font, c.fill, c.alignment, c.border = (
                    Font(name="Arial", size=10), _fill(color), a, _borde())
                if f: c.number_format = f

        # Total
        tr = len(tramos) + 3
        ws3.cell(tr, 1, "TOTAL").font      = Font(name="Arial", bold=True, size=10)
        ws3.cell(tr, 1).fill               = _fill("D9D9D9")
        ws3.cell(tr, 1).alignment          = CENTRO
        ws3.cell(tr, 1).border             = _borde()
        for col in range(2, 6):
            c = ws3.cell(tr, col)
            c.fill, c.border = _fill("D9D9D9"), _borde()
            c.font = Font(name="Arial", bold=True, size=10)
            if col in [3, 4]:
                letra = get_column_letter(col)
                c.value         = f"=SUM({letra}3:{letra}{tr-1})"
                c.number_format = "#,##0"
                c.alignment     = DERECHA
            if col == 5:
                c.value         = f"=D{tr}/C{tr}"
                c.number_format = "0.0%"
                c.alignment     = CENTRO

    # ── Hoja 4: Proveedores ───────────────────────────────────────
    ws4 = wb.create_sheet("Proveedores")
    df_prov = cargar_proveedores(rutas.get("proveedores", ""))
    if df_prov is not None:
        for i, w in enumerate([28, 12, 16, 16, 14], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w

        ws4.merge_cells("A1:E1")
        ws4["A1"] = "RANKING DE PROVEEDORES"
        ws4["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws4["A1"].fill      = _fill(AZUL_OSC)
        ws4["A1"].alignment = CENTRO
        ws4.row_dimensions[1].height = 28

        for col, h in enumerate(["Proveedor", "Facturas", "Monto Total", "Días Pago Prom.", "% Concentración"], 1):
            c = ws4.cell(2, col, h)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", bold=True, color="FFFFFF", size=10),
                _fill(AZUL_OSC), CENTRO, _borde())

        total_prov = df_prov["monto"].sum()
        rank = df_prov.groupby("proveedor").agg(
            facturas   = ("factura",   "count"),
            monto      = ("monto",     "sum"),
            dias_pago  = ("dias_pago", "mean"),
        ).sort_values("monto", ascending=False).reset_index()

        for i, (_, row) in enumerate(rank.iterrows(), 3):
            conc  = row["monto"] / total_prov
            color = "FCE4D6" if conc >= 0.30 else "FFF2CC" if conc >= 0.15 else "FFFFFF"
            color = VERDE if i % 2 == 0 and color == "FFFFFF" else color
            datos = [row["proveedor"], int(row["facturas"]), row["monto"],
                     round(row["dias_pago"]) if pd.notna(row["dias_pago"]) else "—", conc]
            fmts  = [None, "#,##0", "#,##0", "#,##0", "0.0%"]
            alns  = [IZQUIERDA, CENTRO, DERECHA, CENTRO, CENTRO]
            for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
                c = ws4.cell(i, col, v)
                c.font, c.fill, c.alignment, c.border = (
                    Font(name="Arial", size=10), _fill(color), a, _borde())
                if f and isinstance(v, (int, float)): c.number_format = f

    # ── Hoja 5: Líneas de crédito ─────────────────────────────────
    ws5 = wb.create_sheet("Líneas Crédito")
    if df_lin is not None:
        for i, w in enumerate([22, 24, 16, 16, 16, 10], 1):
            ws5.column_dimensions[get_column_letter(i)].width = w

        ws5.merge_cells("A1:F1")
        ws5["A1"] = "LÍNEAS DE CRÉDITO"
        ws5["A1"].font      = Font(name="Arial", bold=True, size=13, color="FFFFFF")
        ws5["A1"].fill      = _fill(AZUL_OSC)
        ws5["A1"].alignment = CENTRO
        ws5.row_dimensions[1].height = 28

        for col, h in enumerate(["Banco", "Tipo", "Cupo Total", "Cupo Usado", "Disponible", "% Uso"], 1):
            c = ws5.cell(2, col, h)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", bold=True, color="FFFFFF", size=10),
                _fill(AZUL_OSC), CENTRO, _borde())

        for i, (_, row) in enumerate(df_lin.iterrows(), 3):
            pct   = row["pct_uso"] if pd.notna(row["pct_uso"]) else 0
            color = ROJO if pct >= 0.95 else "FCE4D6" if pct >= 0.80 else AMARILLO if pct >= 0.50 else VERDE
            datos = [row["banco"], row["tipo_linea"], row["cupo_total"],
                     row["cupo_usado"], row["cupo_disponible"], pct]
            fmts  = [None, None, "#,##0", "#,##0", "#,##0", "0.0%"]
            alns  = [IZQUIERDA, IZQUIERDA, DERECHA, DERECHA, DERECHA, CENTRO]
            for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
                c = ws5.cell(i, col, v)
                c.font, c.fill, c.alignment, c.border = (
                    Font(name="Arial", size=10), _fill(color), a, _borde())
                if f: c.number_format = f

    wb.save(archivo)
    return archivo


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== REPORTE EJECUTIVO ===\n")

    empresa = input("  Nombre empresa          : ").strip() or "Mi Empresa"
    periodo = input("  Período (ej: Marzo 2025): ").strip() or date.today().strftime("%B %Y")

    print("\n  Rutas de archivos (Enter para omitir módulo):\n")
    rutas = {
        "cobranzas":   input("  Cobranzas (.xlsx)        : ").strip(),
        "proveedores": input("  Proveedores (.xlsx)      : ").strip(),
        "lineas":      input("  Líneas de crédito (.xlsx): ").strip(),
        "indicadores": input("  Balance y resultados (.xlsx): ").strip(),
    }

    print(f"\n  Generando reporte ejecutivo...\n")
    archivo = f"reporte_ejecutivo_{date.today().strftime('%Y%m%d')}.xlsx"
    salida  = generar_reporte(empresa, periodo, rutas, archivo)
    print(f"  ✅ Reporte generado: {salida}\n")