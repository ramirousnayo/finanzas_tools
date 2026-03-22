"""
Estimador de Cobranzas y Recaudación
Usa el archivo de cobranzas existente y aplica probabilidades
de cobro por tramo de vencimiento para proyectar la recaudación.

Columnas requeridas (mismo formato que aging):
- cliente, factura, fecha_venc, monto
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
import calendar


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center")
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")

# Probabilidades de cobro por tramo — ajustables
PROBABILIDADES = {
    "Al día":  0.95,
    "0-30":    0.85,
    "31-60":   0.65,
    "61-90":   0.40,
    "+90":     0.15,
}

COLORES = {
    "Al día":  "E2EFDA",
    "0-30":    "FFFFFF",
    "31-60":   "FFF2CC",
    "61-90":   "FCE4D6",
    "+90":     "FFCCCC",
}

COLORES_HEADER = {
    "Al día":  "375623",
    "0-30":    "1F3864",
    "31-60":   "7F6000",
    "61-90":   "843C0C",
    "+90":     "C00000",
}


# ── Lógica ─────────────────────────────────────────────────────────

def clasificar_tramo(dias):
    if dias <= 0:   return "Al día"
    elif dias <= 30: return "0-30"
    elif dias <= 60: return "31-60"
    elif dias <= 90: return "61-90"
    else:            return "+90"


def cargar_y_procesar(ruta):
    try:
        df = pd.read_excel(ruta)
    except FileNotFoundError:
        print(f"\n  ❌ Archivo no encontrado: {ruta}\n")
        return None

    df.columns = df.columns.str.strip().str.lower()

    for col in ["cliente", "factura", "fecha_venc", "monto"]:
        if col not in df.columns:
            print(f"\n  ❌ Falta la columna '{col}'.\n")
            return None

    df["fecha_venc"]    = pd.to_datetime(df["fecha_venc"], dayfirst=True)
    hoy                 = pd.Timestamp(date.today())
    df["dias_vencido"]  = (hoy - df["fecha_venc"]).dt.days
    df["tramo"]         = df["dias_vencido"].apply(clasificar_tramo)
    df["probabilidad"]  = df["tramo"].map(PROBABILIDADES)
    df["monto_estimado"]= df["monto"] * df["probabilidad"]

    return df


def proyeccion_mensual(df):
    """Agrupa la recaudación estimada por mes de vencimiento."""
    df["mes_venc"] = df["fecha_venc"].dt.to_period("M")
    resumen = df.groupby("mes_venc").agg(
        facturas        = ("factura",        "count"),
        monto_total     = ("monto",          "sum"),
        monto_estimado  = ("monto_estimado", "sum"),
    ).reset_index()
    resumen["pct_recuperacion"] = resumen["monto_estimado"] / resumen["monto_total"]
    return resumen


def resumen_por_cliente(df):
    """Agrupa por cliente con probabilidad promedio ponderada."""
    resumen = df.groupby("cliente").agg(
        facturas        = ("factura",        "count"),
        monto_total     = ("monto",          "sum"),
        monto_estimado  = ("monto_estimado", "sum"),
    ).reset_index()
    resumen["prob_promedio"] = resumen["monto_estimado"] / resumen["monto_total"]
    return resumen.sort_values("monto_estimado", ascending=False)


# ── Excel ──────────────────────────────────────────────────────────

def generar_excel(df, df_mensual, df_clientes, archivo):
    wb = Workbook()

    # ── Hoja 1: Detalle con estimación ──
    ws = wb.active
    ws.title = "Detalle"

    anchos = [28, 12, 14, 12, 12, 14, 14]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(anchos))}1")
    ws["A1"] = f"ESTIMADOR DE COBRANZAS — {date.today().strftime('%d/%m/%Y')}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 28

    total_cartera  = df["monto"].sum()
    total_estimado = df["monto_estimado"].sum()
    pct_global     = total_estimado / total_cartera if total_cartera > 0 else 0

    ws.merge_cells(f"A2:{get_column_letter(len(anchos))}2")
    ws["A2"] = (f"Cartera total: ${total_cartera:,.0f}   |   "
                f"Estimado a recaudar: ${total_estimado:,.0f}   |   "
                f"% Recuperación: {pct_global:.1%}")
    ws["A2"].font      = Font(name="Arial", size=9, color="555555")
    ws["A2"].alignment = CENTRO

    headers = ["Cliente", "Factura", "Fecha Venc.", "Días Venc.",
               "Tramo", "Probabilidad", "Monto", "Monto Estimado"]
    anchos2 = [28, 12, 14, 12, 10, 13, 14, 14]
    for i, (h, w) in enumerate(zip(headers, anchos2), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(3, i, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    df_sorted = df.sort_values(["tramo", "monto_estimado"], ascending=[True, False])
    tramos_orden = ["Al día", "0-30", "31-60", "61-90", "+90"]
    df_sorted["tramo"] = pd.Categorical(df_sorted["tramo"], categories=tramos_orden, ordered=True)
    df_sorted = df_sorted.sort_values("tramo")

    for i, (_, row) in enumerate(df_sorted.iterrows(), 1):
        fila    = i + 3
        tramo   = row["tramo"]
        relleno = _fill(COLORES.get(tramo, "FFFFFF"))
        datos   = [
            row["cliente"],
            str(row["factura"]),
            row["fecha_venc"].strftime("%d/%m/%Y"),
            int(row["dias_vencido"]),
            tramo,
            row["probabilidad"],
            row["monto"],
            row["monto_estimado"],
        ]
        fmts = [None, None, None, "#,##0", None, "0%", "#,##0", "#,##0"]
        alns = [IZQUIERDA, CENTRO, CENTRO, CENTRO, CENTRO, CENTRO, DERECHA, DERECHA]

        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws.cell(fila, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), relleno, a, _borde())
            if f:
                c.number_format = f

    # Totales
    tr = len(df) + 4
    ws.cell(tr, 1, "TOTAL").font      = Font(name="Arial", bold=True, size=10)
    ws.cell(tr, 1).fill      = _fill("D9D9D9")
    ws.cell(tr, 1).alignment = CENTRO
    ws.cell(tr, 1).border    = _borde()
    for col in range(2, 9):
        c = ws.cell(tr, col)
        c.fill, c.border = _fill("D9D9D9"), _borde()
        c.font = Font(name="Arial", bold=True, size=10)
        if col in [7, 8]:
            letra           = get_column_letter(col)
            c.value         = f"=SUM({letra}4:{letra}{tr-1})"
            c.number_format = "#,##0"
            c.alignment     = DERECHA

    ws.freeze_panes = "A4"

    # ── Hoja 2: Proyección mensual ──
    ws2 = wb.create_sheet("Proyección Mensual")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "PROYECCIÓN DE RECAUDACIÓN MENSUAL"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws2["A1"].alignment = CENTRO
    ws2.row_dimensions[1].height = 28

    headers2 = ["Mes", "Facturas", "Monto Cartera", "Monto Estimado", "% Recuperación"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(2, i, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    for i, (_, row) in enumerate(df_mensual.iterrows(), 3):
        pct     = row["pct_recuperacion"]
        color   = "E2EFDA" if pct >= 0.7 else "FFF2CC" if pct >= 0.4 else "FFCCCC"
        datos   = [str(row["mes_venc"]), int(row["facturas"]),
                   row["monto_total"], row["monto_estimado"], pct]
        fmts    = [None, "#,##0", "#,##0", "#,##0", "0.0%"]
        alns    = [CENTRO, CENTRO, DERECHA, DERECHA, CENTRO]
        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws2.cell(i, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), _fill(color), a, _borde())
            if f:
                c.number_format = f

    # ── Hoja 3: Resumen por cliente ──
    ws3 = wb.create_sheet("Por Cliente")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 16
    ws3.column_dimensions["E"].width = 16

    ws3.merge_cells("A1:E1")
    ws3["A1"] = "ESTIMACIÓN DE RECAUDACIÓN POR CLIENTE"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws3["A1"].alignment = CENTRO
    ws3.row_dimensions[1].height = 28

    headers3 = ["Cliente", "Facturas", "Monto Cartera", "Monto Estimado", "Prob. Promedio"]
    for i, h in enumerate(headers3, 1):
        c = ws3.cell(2, i, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    for i, (_, row) in enumerate(df_clientes.iterrows(), 3):
        prob    = row["prob_promedio"]
        color   = "E2EFDA" if prob >= 0.7 else "FFF2CC" if prob >= 0.4 else "FFCCCC"
        datos   = [row["cliente"], int(row["facturas"]),
                   row["monto_total"], row["monto_estimado"], prob]
        fmts    = [None, "#,##0", "#,##0", "#,##0", "0.0%"]
        alns    = [IZQUIERDA, CENTRO, DERECHA, DERECHA, CENTRO]
        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws3.cell(i, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), _fill(color), a, _borde())
            if f:
                c.number_format = f

    # ── Hoja 4: Probabilidades usadas ──
    ws4 = wb.create_sheet("Probabilidades")
    ws4.column_dimensions["A"].width = 14
    ws4.column_dimensions["B"].width = 18
    ws4.column_dimensions["C"].width = 30

    ws4.merge_cells("A1:C1")
    ws4["A1"] = "PROBABILIDADES DE COBRO POR TRAMO"
    ws4["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws4["A1"].alignment = CENTRO

    for i, h in enumerate(["Tramo", "Probabilidad", "Interpretación"], 1):
        c = ws4.cell(2, i, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    interpretaciones = {
        "Al día":  "Alta probabilidad — cliente al corriente",
        "0-30":    "Buena probabilidad — atraso leve",
        "31-60":   "Riesgo moderado — seguimiento requerido",
        "61-90":   "Riesgo alto — gestión activa necesaria",
        "+90":     "Riesgo crítico — posible incobrable",
    }

    for i, (tramo, prob) in enumerate(PROBABILIDADES.items(), 3):
        color = COLORES.get(tramo, "FFFFFF")
        ws4.cell(i, 1, tramo).font       = Font(name="Arial", bold=True, size=10)
        ws4.cell(i, 1).fill              = _fill(color)
        ws4.cell(i, 1).alignment         = CENTRO
        ws4.cell(i, 1).border            = _borde()
        c = ws4.cell(i, 2, prob)
        c.font, c.fill, c.alignment, c.border = (
            Font(name="Arial", bold=True, size=10), _fill(color), CENTRO, _borde())
        c.number_format = "0%"
        ws4.cell(i, 3, interpretaciones[tramo]).font      = Font(name="Arial", size=10)
        ws4.cell(i, 3).fill      = _fill(color)
        ws4.cell(i, 3).alignment = IZQUIERDA
        ws4.cell(i, 3).border    = _borde()

    wb.save(archivo)
    return archivo


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== ESTIMADOR DE COBRANZAS Y RECAUDACIÓN ===\n")

    ruta = input("  Ruta archivo de cobranzas (.xlsx): ").strip()

    df = cargar_y_procesar(ruta)
    if df is None:
        return

    print(f"\n  Registros cargados   : {len(df)}")
    print(f"  Cartera total        : ${df['monto'].sum():,.0f}")
    print(f"  Estimado a recaudar  : ${df['monto_estimado'].sum():,.0f}")
    print(f"  % Recuperación       : {df['monto_estimado'].sum() / df['monto'].sum():.1%}\n")

    print("  Probabilidades aplicadas:")
    for tramo, prob in PROBABILIDADES.items():
        subset = df[df["tramo"] == tramo]
        if len(subset) > 0:
            print(f"  {tramo:8} ({prob:.0%})  →  ${subset['monto_estimado'].sum():>14,.0f}")

    df_mensual  = proyeccion_mensual(df)
    df_clientes = resumen_por_cliente(df)

    archivo = f"estimacion_cobranzas_{date.today().strftime('%Y%m%d')}.xlsx"
    salida  = generar_excel(df, df_mensual, df_clientes, archivo)
    print(f"\n  ✅ Excel generado: {salida}\n")