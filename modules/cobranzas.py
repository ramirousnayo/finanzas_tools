"""
Aging de Cobranzas
Lee un Excel exportado del ERP y genera reporte por tramos de vencimiento.

Columnas requeridas en el archivo fuente:
- cliente      : nombre o RUT del cliente
- factura      : número de factura
- fecha_venc   : fecha de vencimiento (DD/MM/YYYY)
- monto        : monto adeudado
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO  = Alignment(horizontal="center", vertical="center")
DERECHA = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left", vertical="center")

# Colores por tramo
COLORES = {
    "Al día":    "E2EFDA",   # verde
    "0-30":      "FFFFFF",   # blanco
    "31-60":     "FFF2CC",   # amarillo
    "61-90":     "FCE4D6",   # naranjo
    "+90":       "FFCCCC",   # rojo claro
}

COLORES_HEADER = {
    "Al día":    "375623",
    "0-30":      "1F3864",
    "31-60":     "7F6000",
    "61-90":     "843C0C",
    "+90":       "C00000",
}


# ── Lógica ─────────────────────────────────────────────────────────

def clasificar_tramo(dias):
    if dias <= 0:
        return "Al día"
    elif dias <= 30:
        return "0-30"
    elif dias <= 60:
        return "31-60"
    elif dias <= 90:
        return "61-90"
    else:
        return "+90"


def procesar_archivo(ruta):
    try:
        df = pd.read_excel(ruta)
    except FileNotFoundError:
        print(f"\n  ❌ Archivo no encontrado: {ruta}\n")
        return None

    # Normalizar columnas a minúsculas
    df.columns = df.columns.str.strip().str.lower()

    columnas_req = ["cliente", "factura", "fecha_venc", "monto"]
    for col in columnas_req:
        if col not in df.columns:
            print(f"\n  ❌ Falta la columna '{col}' en el archivo.\n")
            print(f"  Columnas encontradas: {list(df.columns)}\n")
            return None

    df["fecha_venc"] = pd.to_datetime(df["fecha_venc"], dayfirst=True)
    hoy = pd.Timestamp(date.today())
    df["dias_vencido"] = (hoy - df["fecha_venc"]).dt.days
    df["tramo"] = df["dias_vencido"].apply(clasificar_tramo)

    return df


# ── Excel ──────────────────────────────────────────────────────────

def generar_excel(df, archivo_salida):
    wb = Workbook()

    # ── Hoja 1: Detalle completo ──
    ws = wb.active
    ws.title = "Detalle"

    hoy = date.today().strftime("%d/%m/%Y")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"AGING DE COBRANZAS — {hoy}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO

    ws.merge_cells("A2:F2")
    total_general = df["monto"].sum()
    ws["A2"] = f"Total adeudado: ${total_general:,.0f}   |   Facturas: {len(df)}   |   Clientes: {df['cliente'].nunique()}"
    ws["A2"].font      = Font(name="Arial", size=9, color="555555")
    ws["A2"].alignment = CENTRO

    # Encabezados
    headers = ["Cliente", "Factura", "Fecha Venc.", "Días Vencido", "Tramo", "Monto"]
    anchos  = [28, 14, 14, 14, 10, 16]
    for i, (h, w) in enumerate(zip(headers, anchos), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(3, i, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    # Ordenar por días vencido descendente
    df_sorted = df.sort_values("dias_vencido", ascending=False)

    for i, (_, row) in enumerate(df_sorted.iterrows(), 1):
        fila    = i + 3
        tramo   = row["tramo"]
        relleno = _fill(COLORES.get(tramo, "FFFFFF"))
        datos   = [row["cliente"], str(row["factura"]),
                   row["fecha_venc"].strftime("%d/%m/%Y"),
                   int(row["dias_vencido"]), tramo, row["monto"]]
        fmts    = [None, None, None, "#,##0", None, "#,##0"]
        alns    = [IZQUIERDA, CENTRO, CENTRO, CENTRO, CENTRO, DERECHA]

        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws.cell(fila, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), relleno, a, _borde())
            if f:
                c.number_format = f

    # Fila totales
    tr = len(df) + 4
    ws.merge_cells(f"A{tr}:E{tr}")
    ws[f"A{tr}"] = "TOTAL"
    ws[f"A{tr}"].font      = Font(name="Arial", bold=True, size=10)
    ws[f"A{tr}"].fill      = _fill("D9D9D9")
    ws[f"A{tr}"].alignment = CENTRO
    ws[f"A{tr}"].border    = _borde()
    c = ws.cell(tr, 6)
    c.value          = f"=SUM(F4:F{tr-1})"
    c.font           = Font(name="Arial", bold=True, size=10)
    c.fill           = _fill("D9D9D9")
    c.alignment      = DERECHA
    c.border         = _borde()
    c.number_format  = "#,##0"

    ws.freeze_panes = "A4"

    # ── Hoja 2: Resumen por tramo ──
    ws2 = wb.create_sheet("Resumen por Tramo")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 14

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "RESUMEN POR TRAMO DE VENCIMIENTO"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws2["A1"].alignment = CENTRO

    headers2 = ["Tramo", "Facturas", "Monto", "% del Total"]
    for i, h in enumerate(headers2, 1):
        c = ws2.cell(3, i, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    tramos_orden = ["Al día", "0-30", "31-60", "61-90", "+90"]
    total = df["monto"].sum()

    for i, tramo in enumerate(tramos_orden, 1):
        fila       = i + 3
        subset     = df[df["tramo"] == tramo]
        cant       = len(subset)
        monto      = subset["monto"].sum()
        pct        = monto / total if total > 0 else 0
        relleno    = _fill(COLORES.get(tramo, "FFFFFF"))
        color_txt  = COLORES_HEADER.get(tramo, "000000")

        datos = [tramo, cant, monto, pct]
        fmts  = [None, "#,##0", "#,##0", "0.0%"]
        alns  = [CENTRO, CENTRO, DERECHA, CENTRO]

        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws2.cell(fila, col, v)
            c.font      = Font(name="Arial", bold=True, size=10, color=color_txt)
            c.fill      = relleno
            c.alignment = a
            c.border    = _borde()
            if f:
                c.number_format = f

    # Total resumen
    tr2 = len(tramos_orden) + 4
    for col in range(1, 5):
        c = ws2.cell(tr2, col)
        c.fill, c.border = _fill("D9D9D9"), _borde()
        c.font = Font(name="Arial", bold=True, size=10)
        if col == 1:
            c.value, c.alignment = "TOTAL", CENTRO
        elif col == 2:
            c.value          = f"=SUM(B4:B{tr2-1})"
            c.number_format  = "#,##0"
            c.alignment      = CENTRO
        elif col == 3:
            c.value          = f"=SUM(C4:C{tr2-1})"
            c.number_format  = "#,##0"
            c.alignment      = DERECHA
        elif col == 4:
            c.value, c.alignment = "100%", CENTRO

    wb.save(archivo_salida)
    return archivo_salida


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== AGING DE COBRANZAS ===\n")

    ruta = input("  Ruta del archivo Excel (.xlsx): ").strip()
    if not ruta:
        print("\n  ❌ Debes ingresar una ruta.\n")
        return

    df = procesar_archivo(ruta)
    if df is None:
        return

    # Resumen en consola
    print(f"\n  Registros cargados : {len(df)}")
    print(f"  Clientes únicos    : {df['cliente'].nunique()}")
    print(f"  Total adeudado     : ${df['monto'].sum():,.0f}\n")

    tramos = ["Al día", "0-30", "31-60", "61-90", "+90"]
    alertas = {"61-90": "⚠️ ", "+90": "🔴"}
    for tramo in tramos:
        subset = df[df["tramo"] == tramo]
        if len(subset) > 0:
            icono = alertas.get(tramo, "   ")
            print(f"  {icono} {tramo:8} : {len(subset):>4} facturas   ${subset['monto'].sum():>15,.0f}")

    print()
    archivo = f"aging_cobranzas_{date.today().strftime('%Y%m%d')}.xlsx"
    salida  = generar_excel(df, archivo)
    print(f"  ✅ Excel generado: {salida}\n")