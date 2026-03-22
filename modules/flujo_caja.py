"""
Flujo de Caja Mensual
Genera una plantilla Excel para ingresar datos y luego procesa el reporte.
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center")
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")

CATEGORIAS = {
    "INGRESOS": [
        "Cobranza clientes",
        "Otros ingresos",
    ],
    "EGRESOS": [
        "Pago proveedores",
        "Sueldos y remuneraciones",
        "Préstamos y cuotas",
        "Otros egresos fijos",
    ]
}

COLORES_CAT = {
    "INGRESOS": "E2EFDA",
    "EGRESOS":  "FCE4D6",
}

COLORES_HEADER_CAT = {
    "INGRESOS": "375623",
    "EGRESOS":  "843C0C",
}


# ── Plantilla ──────────────────────────────────────────────────────

def generar_plantilla(meses, anio, saldo_inicial, moneda, archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Flujo de Caja"

    # Encabezado título
    total_cols = meses + 3
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    ws["A1"] = f"FLUJO DE CAJA PROYECTADO {anio} — {moneda}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 28

    # Encabezados columnas
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    meses_nombres = ["Ene","Feb","Mar","Abr","May","Jun",
                     "Jul","Ago","Sep","Oct","Nov","Dic"]

    headers = ["Concepto", "Categoría"] + [meses_nombres[i] for i in range(meses)] + ["TOTAL"]
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = 14 if col > 2 else ws.column_dimensions[get_column_letter(col)].width
        c = ws.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    # Saldo inicial
    ws.cell(3, 1, "Saldo inicial").font      = Font(name="Arial", bold=True, size=10)
    ws.cell(3, 1).fill      = _fill("D6E4F0")
    ws.cell(3, 1).alignment = IZQUIERDA
    ws.cell(3, 1).border    = _borde()
    ws.cell(3, 2, "").border = _borde()
    ws.cell(3, 2).fill      = _fill("D6E4F0")

    for col in range(3, meses + 3):
        c = ws.cell(3, col)
        c.value          = saldo_inicial if col == 3 else ""
        c.font           = Font(name="Arial", bold=True, size=10, color="0000FF")
        c.fill           = _fill("D6E4F0")
        c.alignment      = DERECHA
        c.border         = _borde()
        c.number_format  = "#,##0"

    fila = 4
    filas_ingreso = []
    filas_egreso  = []

    for categoria, conceptos in CATEGORIAS.items():
        # Fila categoría
        ws.merge_cells(f"A{fila}:B{fila}")
        c = ws.cell(fila, 1, categoria)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill(COLORES_HEADER_CAT[categoria])
        c.alignment = IZQUIERDA
        c.border    = _borde()
        for col in range(3, meses + 4):
            ws.cell(fila, col).fill   = _fill(COLORES_HEADER_CAT[categoria])
            ws.cell(fila, col).border = _borde()
        fila += 1

        for concepto in conceptos:
            ws.cell(fila, 1, concepto).font      = Font(name="Arial", size=10)
            ws.cell(fila, 1).fill      = _fill(COLORES_CAT[categoria])
            ws.cell(fila, 1).alignment = IZQUIERDA
            ws.cell(fila, 1).border    = _borde()
            ws.cell(fila, 2, categoria).font     = Font(name="Arial", size=9, color="888888")
            ws.cell(fila, 2).fill      = _fill(COLORES_CAT[categoria])
            ws.cell(fila, 2).alignment = CENTRO
            ws.cell(fila, 2).border    = _borde()

            for col in range(3, meses + 3):
                c = ws.cell(fila, col, 0)
                c.font          = Font(name="Arial", size=10, color="0000FF")
                c.fill          = _fill(COLORES_CAT[categoria])
                c.alignment     = DERECHA
                c.border        = _borde()
                c.number_format = "#,##0"

            # Total fila
            col_total = meses + 3
            inicio    = get_column_letter(3)
            fin       = get_column_letter(meses + 2)
            c = ws.cell(fila, col_total)
            c.value         = f"=SUM({inicio}{fila}:{fin}{fila})"
            c.font          = Font(name="Arial", bold=True, size=10)
            c.fill          = _fill(COLORES_CAT[categoria])
            c.alignment     = DERECHA
            c.border        = _borde()
            c.number_format = "#,##0"

            if categoria == "INGRESOS":
                filas_ingreso.append(fila)
            else:
                filas_egreso.append(fila)
            fila += 1

        # Subtotal categoría
        refs = "+".join([f"{get_column_letter(col)}{f}" for f in (filas_ingreso if categoria == "INGRESOS" else filas_egreso)])
        sub_color = "C6EFCE" if categoria == "INGRESOS" else "FFCCCC"
        ws.cell(fila, 1, f"Total {categoria.capitalize()}").font      = Font(name="Arial", bold=True, size=10)
        ws.cell(fila, 1).fill      = _fill(sub_color)
        ws.cell(fila, 1).alignment = IZQUIERDA
        ws.cell(fila, 1).border    = _borde()
        ws.cell(fila, 2).fill      = _fill(sub_color)
        ws.cell(fila, 2).border    = _borde()

        for col in range(3, meses + 4):
            col_letra = get_column_letter(col)
            filas_ref = filas_ingreso if categoria == "INGRESOS" else filas_egreso
            refs_col  = "+".join([f"{col_letra}{f}" for f in filas_ref])
            c = ws.cell(fila, col)
            c.value         = f"={refs_col}"
            c.font          = Font(name="Arial", bold=True, size=10)
            c.fill          = _fill(sub_color)
            c.alignment     = DERECHA
            c.border        = _borde()
            c.number_format = "#,##0"

        if categoria == "INGRESOS":
            fila_total_ing = fila
        else:
            fila_total_egr = fila
        fila += 1

    # Flujo neto
    ws.cell(fila, 1, "FLUJO NETO").font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws.cell(fila, 1).fill      = _fill("1F3864")
    ws.cell(fila, 1).alignment = IZQUIERDA
    ws.cell(fila, 1).border    = _borde()
    ws.cell(fila, 2).fill      = _fill("1F3864")
    ws.cell(fila, 2).border    = _borde()
    for col in range(3, meses + 4):
        col_letra = get_column_letter(col)
        c = ws.cell(fila, col)
        c.value         = f"={col_letra}{fila_total_ing}-{col_letra}{fila_total_egr}"
        c.font          = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill          = _fill("1F3864")
        c.alignment     = DERECHA
        c.border        = _borde()
        c.number_format = "#,##0"
    fila_neto = fila
    fila += 1

    # Saldo final
    ws.cell(fila, 1, "SALDO FINAL").font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws.cell(fila, 1).fill      = _fill("1F3864")
    ws.cell(fila, 1).alignment = IZQUIERDA
    ws.cell(fila, 1).border    = _borde()
    ws.cell(fila, 2).fill      = _fill("1F3864")
    ws.cell(fila, 2).border    = _borde()
    for col in range(3, meses + 3):
        col_letra     = get_column_letter(col)
        col_ant_letra = get_column_letter(col - 1)
        c = ws.cell(fila, col)
        if col == 3:
            c.value = f"=C3+{col_letra}{fila_neto}"
        else:
            c.value = f"={col_ant_letra}{fila}+{col_letra}{fila_neto}"
        c.font          = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill          = _fill("1F3864")
        c.alignment     = DERECHA
        c.border        = _borde()
        c.number_format = "#,##0"

    # Total saldo final
    ws.cell(fila, meses + 3).fill   = _fill("1F3864")
    ws.cell(fila, meses + 3).border = _borde()

    ws.freeze_panes = "C3"
    wb.save(archivo)
    return archivo


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== FLUJO DE CAJA MENSUAL ===\n")
    print("  [1] Generar plantilla")
    print("  [2] Procesar plantilla completada\n")

    opcion = input("  Selecciona una opción [1-2]: ").strip()

    if opcion == "1":
        anio          = input("  Año del flujo           : ") or str(date.today().year)
        meses         = int(input("  Número de meses (1-12)  : ") or 12)
        saldo_inicial = float(input("  Saldo inicial ($)       : ").replace(",", ""))
        moneda        = input("  Moneda (CLP/USD)        : ") or "CLP"

        archivo = f"flujo_caja_{anio}.xlsx"
        generar_plantilla(meses, int(anio), saldo_inicial, moneda, archivo)
        print(f"\n  ✅ Plantilla generada: {archivo}")
        print(f"  👉 Completa los valores en Excel y vuelve a ejecutar la opción 2.\n")

    elif opcion == "2":
        ruta = input("  Ruta del archivo completado: ").strip()
        try:
            wb = load_workbook(ruta, data_only=True)
            ws = wb.active
            print(f"\n  ✅ Archivo cargado: {ruta}")

            # Buscar filas con saldo negativo
            alertas = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and "SALDO FINAL" in str(row[0]):
                    for i, val in enumerate(row[2:], 1):
                        if isinstance(val, (int, float)) and val < 0:
                            alertas.append((i, val))

            if alertas:
                print("\n  ⚠️  ALERTAS — Meses con saldo negativo proyectado:")
                meses_nombres = ["Ene","Feb","Mar","Abr","May","Jun",
                                 "Jul","Ago","Sep","Oct","Nov","Dic"]
                for mes, saldo in alertas:
                    print(f"     🔴 {meses_nombres[mes-1]}: ${saldo:,.0f}")
            else:
                print("\n  ✅ Sin alertas — todos los meses con saldo positivo.")
            print()

        except FileNotFoundError:
            print(f"\n  ❌ Archivo no encontrado: {ruta}\n")

    else:
        print("\n  Opción no válida.\n")