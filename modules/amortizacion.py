from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from dateutil.relativedelta import relativedelta


# ── Estilos y Constantes ──────────────────────────────────────────

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

CENTRO  = Alignment(horizontal="center")
DERECHA = Alignment(horizontal="right")

# Colores y fuentes
COLOR_TITULO = "1F3864"
COLOR_TEXTO_ENCABEZADO = "FFFFFF"
COLOR_FONDO_ENCABEZADO = "1F3864"
COLOR_OC_FILL = "FFF2CC"
COLOR_OC_FONT = "7F4F00"
COLOR_FILA_PAR = "D6E4F0"
COLOR_FILA_IMPAR = "FFFFFF"
COLOR_TOTAL_FILL = "D9D9D9"


# ── Cálculos ───────────────────────────────────────────────────────

def cuota_frances(capital, tasa_mensual, plazo):
    """Calcula la cuota fija del sistema francés."""
    if tasa_mensual == 0:
        return capital / plazo
    return capital * tasa_mensual / (1 - (1 + tasa_mensual) ** -plazo)


def filas_frances(capital, tasa_mensual, plazo, fecha_inicio):
    cuota = cuota_frances(capital, tasa_mensual, plazo)
    filas, saldo, fecha = [], capital, fecha_inicio
    for mes in range(1, plazo + 1):
        interes = saldo * tasa_mensual
        cap     = cuota - interes
        if mes == plazo:
            cap     = saldo
            cuota_r = cap + interes
        else:
            cuota_r = cuota
        s_fin = max(saldo - cap, 0)
        filas.append((mes, fecha, saldo, cuota_r, interes, cap, s_fin))
        saldo  = s_fin
        fecha += relativedelta(months=1)
    return filas, cuota


def filas_aleman(capital, tasa_mensual, plazo, fecha_inicio):
    """Cuota de capital fija — Sistema Alemán."""
    cap_fijo = capital / plazo
    filas, saldo, fecha = [], capital, fecha_inicio
    for mes in range(1, plazo + 1):
        interes = saldo * tasa_mensual
        cuota_r = cap_fijo + interes
        s_fin   = max(saldo - cap_fijo, 0)
        filas.append((mes, fecha, saldo, cuota_r, interes, cap_fijo, s_fin))
        saldo  = s_fin
        fecha += relativedelta(months=1)
    return filas


def filas_americano(capital, tasa_mensual, plazo, fecha_inicio):
    """Intereses periódicos + capital al final — Sistema Americano."""
    interes = capital * tasa_mensual
    filas, fecha = [], fecha_inicio
    for mes in range(1, plazo + 1):
        es_fin = (mes == plazo)
        cuota_r = (capital + interes) if es_fin else interes
        cap     = capital if es_fin else 0
        s_fin   = 0 if es_fin else capital
        filas.append((mes, fecha, capital, cuota_r, interes, cap, s_fin))
        fecha += relativedelta(months=1)
    return filas


def filas_bullet(capital, tasa_mensual, plazo, fecha_inicio):
    """Pago único al vencimiento — Sistema Bullet."""
    interes_total = capital * ((1 + tasa_mensual) ** plazo - 1)
    filas, fecha  = [], fecha_inicio
    for mes in range(1, plazo + 1):
        if mes < plazo:
            filas.append((mes, fecha, capital, 0, 0, 0, capital))
        else:
            pago = capital + interes_total
            filas.append((mes, fecha, capital, pago, interes_total, capital, 0))
        fecha += relativedelta(months=1)
    return filas


def filas_leasing(capital, tasa_mensual, plazo, fecha_inicio):
    cuota = cuota_frances(capital, tasa_mensual, plazo)
    filas, saldo, fecha = [], capital, fecha_inicio
    for mes in range(1, plazo + 1):
        interes = saldo * tasa_mensual
        cap     = cuota - interes
        s_fin   = max(saldo - cap, 0)
        filas.append((mes, fecha, saldo, cuota, interes, cap, s_fin))
        saldo  = s_fin
        fecha += relativedelta(months=1)

    # Opción de compra — cuota adicional al final
    filas.append((plazo + 1, fecha, 0, cuota, 0, cuota, 0))
    return filas, cuota


# ── Excel ──────────────────────────────────────────────────────────

def _formatear_celda(celda, valor, font=None, fill=None, alignment=None, border=None, num_format=None):
    if valor is not None:
        try:
            celda.value = valor
        except AttributeError:
            pass # MergedCell
    if font: celda.font = font
    if fill: celda.fill = fill
    if alignment: celda.alignment = alignment
    if border: celda.border = border
    if num_format: celda.number_format = num_format


def generar_excel(filas, capital, tasa, plazo, moneda, nombre, sistema="Francés"):
    wb = Workbook()
    ws = wb.active
    ws.title = sistema[:31] # Límite de openpyxl

    # Estilos comunes
    font_header = Font(name="Arial", bold=True, color=COLOR_TEXTO_ENCABEZADO, size=10)
    font_data   = Font(name="Arial", size=10)
    font_oc     = Font(name="Arial", size=10, bold=True, color=COLOR_OC_FONT)
    font_total  = Font(name="Arial", bold=True, size=10)
    fill_header = _fill(COLOR_FONDO_ENCABEZADO)
    fill_oc     = _fill(COLOR_OC_FILL)
    fill_total  = _fill(COLOR_TOTAL_FILL)
    borde       = _borde()

    # Título
    ws.merge_cells("A1:G1")
    _formatear_celda(ws["A1"], f"TABLA DE AMORTIZACIÓN — {sistema.upper()} — {nombre.upper()}",
                     font=Font(name="Arial", bold=True, size=13, color=COLOR_TITULO), alignment=CENTRO)

    total_cuotas = sum(f[3] for f in filas)
    total_int    = sum(f[4] for f in filas)
    opcion_compra = filas[-1][3] if sistema == "Leasing" else 0
    oc_txt = f"   |   Opción de compra: {moneda} {opcion_compra:,.0f}" if opcion_compra else ""

    ws.merge_cells("A2:G2")
    info_text = (f"Capital: {moneda} {capital:,.0f}   |   Tasa anual: {tasa:.2f}%   |   "
                 f"Plazo: {plazo} meses   |   Total pagado: {moneda} {total_cuotas:,.0f}   |   "
                 f"Total intereses: {moneda} {total_int:,.0f}{oc_txt}")
    _formatear_celda(ws["A2"], info_text, font=Font(name="Arial", size=9, color="555555"), alignment=CENTRO)

    # Encabezados
    cols = [("N°",6), ("Fecha",13), ("Saldo Inicial",17), ("Cuota",14),
            ("Interés",13), ("Capital",13), ("Saldo Final",14)]
    for i, (h, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        _formatear_celda(ws.cell(3, i), h, font=font_header, fill=fill_header, alignment=CENTRO, border=borde)

    # Filas
    for mes, fecha, s_ini, cuota_r, interes, cap, s_fin in filas:
        row   = mes + 3
        es_oc = (sistema == "Leasing" and mes == plazo + 1)
        
        relleno = fill_oc if es_oc else _fill(COLOR_FILA_PAR if mes % 2 == 0 else COLOR_FILA_IMPAR)
        fuente  = font_oc if es_oc else font_data
        
        datos = [mes, fecha, s_ini, cuota_r, interes, cap, s_fin]
        fmts  = ["0", "DD/MM/YYYY", "#,##0", "#,##0", "#,##0", "#,##0", "#,##0"]
        alns  = [CENTRO, CENTRO, DERECHA, DERECHA, DERECHA, DERECHA, DERECHA]
        
        for col, (v, fmt, aln) in enumerate(zip(datos, fmts, alns), 1):
            val = "OC" if es_oc and col == 1 else v
            _formatear_celda(ws.cell(row, col), val, font=fuente, fill=relleno, alignment=aln, border=borde, num_format=fmt)

    # Totales
    tr = len(filas) + 4
    ws.merge_cells(f"A{tr}:B{tr}")
    for col in range(1, 8):
        c = ws.cell(tr, col)
        val = "TOTAL" if col == 1 else (f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{tr-1})" if col in [4, 5, 6] else None)
        _formatear_celda(c, val, font=font_total, fill=fill_total, alignment=CENTRO if col == 1 else DERECHA, border=borde, num_format="#,##0" if col in [4,5,6] else None)

    ws.freeze_panes = "A4"
    archivo = f"amortizacion_{sistema.lower()}_{nombre.replace(' ', '_')}.xlsx"
    wb.save(archivo)
    return archivo


# ── Menú y Orquestación ────────────────────────────────────────────

def _get_input(msg, default=None, parser=str):
    """Helper para obtener input con validación básica."""
    try:
        val = input(f"  {msg:<18}: ").strip().replace(",", "")
        if not val and default is not None: return default
        return parser(val)
    except ValueError:
        print(f"  Entrada inválida. Usando default: {default}")
        return default


def run():
    print("\n=== TABLA DE AMORTIZACIÓN ===\n")
    sistemas = {
        "1": ("Francés",   filas_frances),
        "2": ("Alemán",    filas_aleman),
        "3": ("Americano", filas_americano),
        "4": ("Bullet",    filas_bullet),
        "5": ("Leasing",   filas_leasing)
    }
    
    for k, (name, _) in sistemas.items():
        print(f"  [{k}] {name}")
    print()

    opt = input("  Selecciona sistema [1-5]: ").strip()
    if opt not in sistemas:
        print("  Opción no válida."); return

    name_sys, func = sistemas[opt]
    capital = _get_input("Capital", parser=float)
    tasa    = _get_input("Tasa anual (%)", parser=float)
    plazo   = _get_input("Plazo (meses)", parser=int)
    nombre  = _get_input("Nombre préstamo", default="Prestamo")
    moneda  = _get_input("Moneda (CLP/USD)", default="CLP")

    tasa_mensual = tasa / 100 / 12
    fecha_inicio = date.today().replace(day=1) + relativedelta(months=1)

    res = func(capital, tasa_mensual, plazo, fecha_inicio)
    filas, extra = res if isinstance(res, tuple) else (res, None)

    if opt in ["1", "5"]: print(f"\n  Cuota mensual      : {moneda} {extra:,.0f}")
    elif opt == "2":      print(f"\n  Capital por cuota  : {moneda} {capital/plazo:,.0f}")
    elif opt == "3":      print(f"\n  Interés mensual    : {moneda} {capital * tasa_mensual:,.0f}")
    elif opt == "4":      print(f"\n  Pago único final   : {moneda} {capital * ((1 + tasa_mensual) ** plazo):,.0f}")

    if opt == "5":
        print(f"  Opción de compra   : {moneda} {extra:,.0f}  (cuota {plazo + 1})")
        print(f"\n  ⚠️  Al mes {plazo + 1} deberás pagar la opción de compra.")

    if input("\n  ¿Deseas generar la tabla? [s/n]: ").strip().lower() == "s":
        archivo = generar_excel(filas, capital, tasa, plazo, moneda, nombre, name_sys)
        print(f"  ✅ Excel generado: {archivo}\n")