"""
Análisis de Proveedores
Ranking, concentración, días de pago y proveedores críticos.

Columnas requeridas en el Excel del ERP:
- proveedor    : nombre del proveedor
- rut          : RUT del proveedor
- factura      : número de factura
- fecha_factura: fecha emisión factura (DD/MM/YYYY)
- fecha_pago   : fecha de pago (DD/MM/YYYY) — vacío si no pagado
- monto        : monto de la factura
- credito_dias : días de crédito pactado (default 30)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center")
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")


# ── Lógica ─────────────────────────────────────────────────────────

def cargar_archivo(ruta):
    try:
        df = pd.read_excel(ruta)
    except FileNotFoundError:
        print(f"\n  ❌ Archivo no encontrado: {ruta}\n")
        return None

    df.columns = df.columns.str.strip().str.lower()

    for col in ["proveedor", "rut", "factura", "fecha_factura", "monto"]:
        if col not in df.columns:
            print(f"\n  ❌ Falta la columna '{col}'.")
            print(f"  Columnas encontradas: {list(df.columns)}\n")
            return None

    df["fecha_factura"] = pd.to_datetime(df["fecha_factura"], dayfirst=True)
    df["monto"]         = pd.to_numeric(df["monto"], errors="coerce")
    df["credito_dias"]  = df.get("credito_dias", 30).fillna(30).astype(int)

    # Fecha pago — puede estar vacía (facturas pendientes)
    if "fecha_pago" in df.columns:
        df["fecha_pago"] = pd.to_datetime(df["fecha_pago"], dayfirst=True, errors="coerce")
    else:
        df["fecha_pago"] = pd.NaT

    hoy = pd.Timestamp(date.today())

    # Días reales de pago (solo facturas pagadas)
    df["dias_pago"] = (df["fecha_pago"] - df["fecha_factura"]).dt.days

    # Fecha límite de pago según crédito pactado
    df["fecha_limite"] = df["fecha_factura"] + pd.to_timedelta(df["credito_dias"], unit="d")

    # Estado de cada factura
    def estado(row):
        if pd.notna(row["fecha_pago"]):
            if row["dias_pago"] <= row["credito_dias"]:
                return "Pagado a tiempo"
            else:
                return "Pagado con atraso"
        else:
            if hoy <= row["fecha_limite"]:
                return "Pendiente vigente"
            else:
                return "Pendiente vencido"

    df["estado"] = df.apply(estado, axis=1)

    # Días de atraso para pendientes vencidos
    df["dias_atraso"] = df.apply(
        lambda r: (hoy - r["fecha_limite"]).days
        if r["estado"] == "Pendiente vencido" else 0, axis=1)

    return df


def ranking_proveedores(df):
    total = df["monto"].sum()
    rank  = df.groupby(["proveedor", "rut"]).agg(
        facturas        = ("factura",    "count"),
        monto_total     = ("monto",      "sum"),
        dias_pago_prom  = ("dias_pago",  "mean"),
    ).reset_index()
    rank["concentracion"]   = rank["monto_total"] / total
    rank["concentracion_ac"]= rank["monto_total"].cumsum() / total
    rank = rank.sort_values("monto_total", ascending=False).reset_index(drop=True)
    rank.index += 1
    return rank


def proveedores_criticos(df):
    """Proveedores con facturas pendientes vencidas."""
    criticos = df[df["estado"] == "Pendiente vencido"].copy()
    if criticos.empty:
        return criticos
    return criticos.groupby(["proveedor", "rut"]).agg(
        facturas_vencidas = ("factura",      "count"),
        monto_vencido     = ("monto",        "sum"),
        dias_atraso_max   = ("dias_atraso",  "max"),
        dias_atraso_prom  = ("dias_atraso",  "mean"),
    ).reset_index().sort_values("monto_vencido", ascending=False)


def historial_mensual(df):
    """Monto pagado y pendiente por mes."""
    df["mes"] = df["fecha_factura"].dt.to_period("M")
    return df.groupby(["mes", "estado"]).agg(
        facturas = ("factura", "count"),
        monto    = ("monto",   "sum"),
    ).reset_index().sort_values("mes")


# ── Excel ──────────────────────────────────────────────────────────

def escribir_encabezado(ws, titulo, subtitulo=None, cols=5):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    ws["A1"] = titulo
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 28
    if subtitulo:
        ws.merge_cells(f"A2:{get_column_letter(cols)}2")
        ws["A2"] = subtitulo
        ws["A2"].font      = Font(name="Arial", size=9, color="555555")
        ws["A2"].alignment = CENTRO


def escribir_headers(ws, headers, row=3, color="1F3864"):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill(color)
        c.alignment = CENTRO
        c.border    = _borde()


def generar_excel(df, rank, criticos, historial, archivo):
    wb = Workbook()
    total = df["monto"].sum()
    hoy   = date.today().strftime("%d/%m/%Y")

    # ── Hoja 1: Ranking ──
    ws = wb.active
    ws.title = "Ranking"
    anchos = [28, 14, 10, 16, 16, 14, 14]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    escribir_encabezado(ws, f"RANKING DE PROVEEDORES — {hoy}",
                        f"Total compras: ${total:,.0f}   |   Proveedores: {len(rank)}", 7)
    escribir_headers(ws, ["#", "Proveedor", "RUT", "Facturas",
                           "Monto Total", "Concentración", "Concentración Ac."], color="1F3864")

    for i, (_, row) in enumerate(rank.iterrows(), 1):
        fila    = i + 3
        conc    = row["concentracion"]
        color   = "FCE4D6" if conc >= 0.30 else "FFF2CC" if conc >= 0.15 else "FFFFFF"
        relleno = _fill(color) if i % 2 == 0 else _fill("F5F5F5" if color == "FFFFFF" else color)
        datos   = [i, row["proveedor"], row["rut"], int(row["facturas"]),
                   row["monto_total"], row["concentracion"], row["concentracion_ac"]]
        fmts    = [None, None, None, "#,##0", "#,##0", "0.0%", "0.0%"]
        alns    = [CENTRO, IZQUIERDA, CENTRO, CENTRO, DERECHA, CENTRO, CENTRO]
        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws.cell(fila, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), relleno, a, _borde())
            if f:
                c.number_format = f

    ws.freeze_panes = "A4"

    # ── Hoja 2: Críticos ──
    ws2 = wb.create_sheet("Críticos")
    for i, w in enumerate([28, 14, 12, 16, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    escribir_encabezado(ws2, f"PROVEEDORES CRÍTICOS — FACTURAS VENCIDAS",
                        f"Generado: {hoy}", 6)
    escribir_headers(ws2, ["Proveedor", "RUT", "Fact. Vencidas",
                            "Monto Vencido", "Días Atraso Máx.", "Días Atraso Prom."],
                     color="C00000")

    if criticos.empty:
        ws2.merge_cells("A4:F4")
        ws2["A4"] = "✅ Sin proveedores críticos"
        ws2["A4"].font      = Font(name="Arial", size=10, color="375623")
        ws2["A4"].alignment = CENTRO
    else:
        for i, (_, row) in enumerate(criticos.iterrows(), 4):
            atraso  = row["dias_atraso_max"]
            color   = "FFCCCC" if atraso > 60 else "FCE4D6" if atraso > 30 else "FFF2CC"
            datos   = [row["proveedor"], row["rut"], int(row["facturas_vencidas"]),
                       row["monto_vencido"], int(row["dias_atraso_max"]),
                       round(row["dias_atraso_prom"])]
            fmts    = [None, None, "#,##0", "#,##0", "#,##0", "#,##0"]
            alns    = [IZQUIERDA, CENTRO, CENTRO, DERECHA, CENTRO, CENTRO]
            for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
                c = ws2.cell(i, col, v)
                c.font, c.fill, c.alignment, c.border = (
                    Font(name="Arial", size=10), _fill(color), a, _borde())
                if f:
                    c.number_format = f

    # ── Hoja 3: Historial mensual ──
    ws3 = wb.create_sheet("Historial Mensual")
    for i, w in enumerate([14, 22, 10, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    escribir_encabezado(ws3, "HISTORIAL MENSUAL DE PAGOS", f"Generado: {hoy}", 4)
    escribir_headers(ws3, ["Mes", "Estado", "Facturas", "Monto"], color="1F3864")

    colores_estado = {
        "Pagado a tiempo":    "E2EFDA",
        "Pagado con atraso":  "FFF2CC",
        "Pendiente vigente":  "D6E4F0",
        "Pendiente vencido":  "FFCCCC",
    }

    for i, (_, row) in enumerate(historial.iterrows(), 4):
        color = colores_estado.get(row["estado"], "FFFFFF")
        datos = [str(row["mes"]), row["estado"], int(row["facturas"]), row["monto"]]
        fmts  = [None, None, "#,##0", "#,##0"]
        alns  = [CENTRO, IZQUIERDA, CENTRO, DERECHA]
        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws3.cell(i, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), _fill(color), a, _borde())
            if f:
                c.number_format = f

    # ── Hoja 4: Detalle completo ──
    ws4 = wb.create_sheet("Detalle")
    for i, w in enumerate([28, 14, 12, 14, 14, 14, 10, 22], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    escribir_encabezado(ws4, "DETALLE COMPLETO DE FACTURAS", f"Generado: {hoy}", 8)
    escribir_headers(ws4, ["Proveedor", "RUT", "Factura", "Fecha Factura",
                            "Fecha Pago", "Monto", "Días Pago", "Estado"], color="1F3864")

    df_sorted = df.sort_values(["proveedor", "fecha_factura"])
    for i, (_, row) in enumerate(df_sorted.iterrows(), 4):
        color = colores_estado.get(row["estado"], "FFFFFF")
        fecha_pago = row["fecha_pago"].strftime("%d/%m/%Y") if pd.notna(row["fecha_pago"]) else "—"
        dias_pago  = int(row["dias_pago"]) if pd.notna(row["dias_pago"]) else "—"
        datos = [row["proveedor"], row["rut"], str(row["factura"]),
                 row["fecha_factura"].strftime("%d/%m/%Y"),
                 fecha_pago, row["monto"], dias_pago, row["estado"]]
        fmts  = [None, None, None, None, None, "#,##0", "#,##0" if isinstance(dias_pago, int) else None, None]
        alns  = [IZQUIERDA, CENTRO, CENTRO, CENTRO, CENTRO, DERECHA, CENTRO, CENTRO]
        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws4.cell(i, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), _fill(color), a, _borde())
            if f and isinstance(v, (int, float)):
                c.number_format = f

    ws4.freeze_panes = "A4"

    wb.save(archivo)
    return archivo


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== ANÁLISIS DE PROVEEDORES ===\n")

    ruta = input("  Ruta archivo de proveedores (.xlsx): ").strip()

    df = cargar_archivo(ruta)
    if df is None:
        return

    print(f"\n  Registros cargados  : {len(df)}")
    print(f"  Proveedores únicos  : {df['proveedor'].nunique()}")
    print(f"  Monto total         : ${df['monto'].sum():,.0f}\n")

    estados = df["estado"].value_counts()
    for estado, cant in estados.items():
        monto = df[df["estado"] == estado]["monto"].sum()
        print(f"  {estado:<22} : {cant:>3} facturas   ${monto:>14,.0f}")

    rank     = ranking_proveedores(df)
    criticos = proveedores_criticos(df)
    historial = historial_mensual(df)

    print(f"\n  Top 3 proveedores por monto:")
    for i, (_, row) in enumerate(rank.head(3).iterrows(), 1):
        print(f"  {i}. {row['proveedor']:<25} ${row['monto_total']:>14,.0f}  ({row['concentracion']:.1%})")

    if not criticos.empty:
        print(f"\n  ⚠️  {len(criticos)} proveedores con facturas vencidas")

    archivo = f"analisis_proveedores_{date.today().strftime('%Y%m%d')}.xlsx"
    salida  = generar_excel(df, rank, criticos, historial, archivo)
    print(f"\n  ✅ Excel generado: {salida}\n")