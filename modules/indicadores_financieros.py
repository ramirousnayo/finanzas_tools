"""
Indicadores Financieros
Calcula ratios financieros clave con semáforo de alertas.

Columnas requeridas en el Excel:
Hoja "Balance":
- cuenta, valor_actual, valor_anterior

Hoja "Resultados":
- cuenta, valor_actual, valor_anterior
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center")
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")

# Semáforo
VERDE    = "E2EFDA"
AMARILLO = "FFF2CC"
ROJO     = "FFCCCC"
VERDE_T  = "375623"
AMARILLO_T = "7F6000"
ROJO_T   = "C00000"


# ── Cuentas esperadas ──────────────────────────────────────────────

CUENTAS_BALANCE = [
    "activo_corriente",
    "activo_no_corriente",
    "pasivo_corriente",
    "pasivo_no_corriente",
    "patrimonio",
    "cuentas_por_cobrar",
    "inventario",
    "caja_y_bancos",
    "deuda_financiera",
    "cuentas_por_pagar",
]

CUENTAS_RESULTADOS = [
    "ingresos",
    "costo_ventas",
    "gastos_operacionales",
    "ebitda",
    "depreciacion",
    "ebit",
    "gastos_financieros",
    "utilidad_neta",
]


# ── Lógica ─────────────────────────────────────────────────────────

def cargar_datos(ruta):
    try:
        wb = load_workbook(ruta, data_only=True)
    except FileNotFoundError:
        print(f"\n  ❌ Archivo no encontrado: {ruta}\n")
        return None, None

    if "Balance" not in wb.sheetnames or "Resultados" not in wb.sheetnames:
        print(f"\n  ❌ El archivo debe tener hojas 'Balance' y 'Resultados'.\n")
        return None, None

    df_b = pd.read_excel(ruta, sheet_name="Balance")
    df_r = pd.read_excel(ruta, sheet_name="Resultados")

    df_b.columns = df_b.columns.str.strip().str.lower()
    df_r.columns = df_r.columns.str.strip().str.lower()

    df_b["cuenta"] = df_b["cuenta"].str.strip().str.lower()
    df_r["cuenta"] = df_r["cuenta"].str.strip().str.lower()

    balance    = df_b.set_index("cuenta")["valor_actual"].to_dict()
    resultados = df_r.set_index("cuenta")["valor_actual"].to_dict()

    balance_ant    = df_b.set_index("cuenta")["valor_anterior"].to_dict()
    resultados_ant = df_r.set_index("cuenta")["valor_anterior"].to_dict()

    return (balance, balance_ant), (resultados, resultados_ant)


def get(d, key, default=0):
    return d.get(key, default) or default


def calcular_indicadores(balance, balance_ant, resultados, resultados_ant):
    b  = balance
    ba = balance_ant
    r  = resultados
    ra = resultados_ant

    indicadores = []

    # ── Liquidez ──
    liq_corriente = get(b, "activo_corriente") / get(b, "pasivo_corriente") if get(b, "pasivo_corriente") else 0
    liq_corriente_ant = get(ba, "activo_corriente") / get(ba, "pasivo_corriente") if get(ba, "pasivo_corriente") else 0

    indicadores.append({
        "categoria":   "Liquidez",
        "indicador":   "Liquidez Corriente",
        "formula":     "Activo Corriente / Pasivo Corriente",
        "valor":       liq_corriente,
        "valor_ant":   liq_corriente_ant,
        "formato":     "ratio",
        "meta":        ">= 1.5",
        "semaforo":    "verde" if liq_corriente >= 1.5 else "amarillo" if liq_corriente >= 1.0 else "rojo",
        "interpretacion": "Capacidad de pagar obligaciones de corto plazo",
    })

    # Prueba ácida
    prueba_acida = (get(b, "activo_corriente") - get(b, "inventario")) / get(b, "pasivo_corriente") if get(b, "pasivo_corriente") else 0
    prueba_acida_ant = (get(ba, "activo_corriente") - get(ba, "inventario")) / get(ba, "pasivo_corriente") if get(ba, "pasivo_corriente") else 0

    indicadores.append({
        "categoria":   "Liquidez",
        "indicador":   "Prueba Ácida",
        "formula":     "(Activo Corriente - Inventario) / Pasivo Corriente",
        "valor":       prueba_acida,
        "valor_ant":   prueba_acida_ant,
        "formato":     "ratio",
        "meta":        ">= 1.0",
        "semaforo":    "verde" if prueba_acida >= 1.0 else "amarillo" if prueba_acida >= 0.7 else "rojo",
        "interpretacion": "Liquidez sin considerar inventario",
    })

    # ── Rotación de cartera ──
    rot_cartera = get(r, "ingresos") / get(b, "cuentas_por_cobrar") if get(b, "cuentas_por_cobrar") else 0
    rot_cartera_ant = get(ra, "ingresos") / get(ba, "cuentas_por_cobrar") if get(ba, "cuentas_por_cobrar") else 0

    indicadores.append({
        "categoria":   "Eficiencia",
        "indicador":   "Rotación de Cartera",
        "formula":     "Ingresos / Cuentas por Cobrar",
        "valor":       rot_cartera,
        "valor_ant":   rot_cartera_ant,
        "formato":     "ratio",
        "meta":        ">= 6x",
        "semaforo":    "verde" if rot_cartera >= 6 else "amarillo" if rot_cartera >= 4 else "rojo",
        "interpretacion": "Veces que se cobra la cartera al año",
    })

    # Días de cobro
    dias_cobro = 365 / rot_cartera if rot_cartera else 0
    dias_cobro_ant = 365 / rot_cartera_ant if rot_cartera_ant else 0

    indicadores.append({
        "categoria":   "Eficiencia",
        "indicador":   "Días de Cobro Promedio",
        "formula":     "365 / Rotación de Cartera",
        "valor":       dias_cobro,
        "valor_ant":   dias_cobro_ant,
        "formato":     "dias",
        "meta":        "<= 60 días",
        "semaforo":    "verde" if dias_cobro <= 60 else "amarillo" if dias_cobro <= 90 else "rojo",
        "interpretacion": "Días promedio para cobrar una factura",
    })

    # Días de pago
    rot_pago = get(r, "costo_ventas") / get(b, "cuentas_por_pagar") if get(b, "cuentas_por_pagar") else 0
    rot_pago_ant = get(ra, "costo_ventas") / get(ba, "cuentas_por_pagar") if get(ba, "cuentas_por_pagar") else 0
    dias_pago = 365 / rot_pago if rot_pago else 0
    dias_pago_ant = 365 / rot_pago_ant if rot_pago_ant else 0

    indicadores.append({
        "categoria":   "Eficiencia",
        "indicador":   "Días de Pago Promedio",
        "formula":     "365 / (Costo Ventas / Cuentas por Pagar)",
        "valor":       dias_pago,
        "valor_ant":   dias_pago_ant,
        "formato":     "dias",
        "meta":        "30 - 60 días",
        "semaforo":    "verde" if 30 <= dias_pago <= 60 else "amarillo" if dias_pago <= 90 else "rojo",
        "interpretacion": "Días promedio para pagar a proveedores",
    })

    # ── Endeudamiento ──
    activo_total = get(b, "activo_corriente") + get(b, "activo_no_corriente")
    activo_total_ant = get(ba, "activo_corriente") + get(ba, "activo_no_corriente")

    endeudamiento = (get(b, "pasivo_corriente") + get(b, "pasivo_no_corriente")) / activo_total if activo_total else 0
    endeudamiento_ant = (get(ba, "pasivo_corriente") + get(ba, "pasivo_no_corriente")) / activo_total_ant if activo_total_ant else 0

    indicadores.append({
        "categoria":   "Endeudamiento",
        "indicador":   "Razón de Endeudamiento",
        "formula":     "Pasivo Total / Activo Total",
        "valor":       endeudamiento,
        "valor_ant":   endeudamiento_ant,
        "formato":     "porcentaje",
        "meta":        "<= 60%",
        "semaforo":    "verde" if endeudamiento <= 0.60 else "amarillo" if endeudamiento <= 0.75 else "rojo",
        "interpretacion": "% de activos financiados con deuda",
    })

    # Concentración deuda financiera
    conc_deuda = get(b, "deuda_financiera") / activo_total if activo_total else 0
    conc_deuda_ant = get(ba, "deuda_financiera") / activo_total_ant if activo_total_ant else 0

    indicadores.append({
        "categoria":   "Endeudamiento",
        "indicador":   "Concentración Deuda Financiera",
        "formula":     "Deuda Financiera / Activo Total",
        "valor":       conc_deuda,
        "valor_ant":   conc_deuda_ant,
        "formato":     "porcentaje",
        "meta":        "<= 40%",
        "semaforo":    "verde" if conc_deuda <= 0.40 else "amarillo" if conc_deuda <= 0.55 else "rojo",
        "interpretacion": "% de activos financiados con deuda bancaria",
    })

    # ── Ratios bancarios ──
    ebitda = get(r, "ebitda")
    ebitda_ant = get(ra, "ebitda")

    # Deuda / EBITDA
    deuda_ebitda = get(b, "deuda_financiera") / ebitda if ebitda else 0
    deuda_ebitda_ant = get(ba, "deuda_financiera") / ebitda_ant if ebitda_ant else 0

    indicadores.append({
        "categoria":   "Ratios Bancarios",
        "indicador":   "Deuda Financiera / EBITDA",
        "formula":     "Deuda Financiera / EBITDA",
        "valor":       deuda_ebitda,
        "valor_ant":   deuda_ebitda_ant,
        "formato":     "ratio",
        "meta":        "<= 3.0x",
        "semaforo":    "verde" if deuda_ebitda <= 3.0 else "amarillo" if deuda_ebitda <= 4.5 else "rojo",
        "interpretacion": "Años para pagar deuda con EBITDA — clave para bancos",
    })

    # Cobertura de intereses
    ebit = get(r, "ebit")
    ebit_ant = get(ra, "ebit")
    gf   = get(r, "gastos_financieros")
    gf_ant = get(ra, "gastos_financieros")

    cobertura = ebit / gf if gf else 0
    cobertura_ant = ebit_ant / gf_ant if gf_ant else 0

    indicadores.append({
        "categoria":   "Ratios Bancarios",
        "indicador":   "Cobertura de Intereses",
        "formula":     "EBIT / Gastos Financieros",
        "valor":       cobertura,
        "valor_ant":   cobertura_ant,
        "formato":     "ratio",
        "meta":        ">= 2.5x",
        "semaforo":    "verde" if cobertura >= 2.5 else "amarillo" if cobertura >= 1.5 else "rojo",
        "interpretacion": "Capacidad de pagar intereses con utilidad operacional",
    })

    # Margen EBITDA
    margen_ebitda = ebitda / get(r, "ingresos") if get(r, "ingresos") else 0
    margen_ebitda_ant = ebitda_ant / get(ra, "ingresos") if get(ra, "ingresos") else 0

    indicadores.append({
        "categoria":   "Rentabilidad",
        "indicador":   "Margen EBITDA",
        "formula":     "EBITDA / Ingresos",
        "valor":       margen_ebitda,
        "valor_ant":   margen_ebitda_ant,
        "formato":     "porcentaje",
        "meta":        ">= 15%",
        "semaforo":    "verde" if margen_ebitda >= 0.15 else "amarillo" if margen_ebitda >= 0.08 else "rojo",
        "interpretacion": "% de ingresos que se convierte en EBITDA",
    })

    # Margen neto
    margen_neto = get(r, "utilidad_neta") / get(r, "ingresos") if get(r, "ingresos") else 0
    margen_neto_ant = get(ra, "utilidad_neta") / get(ra, "ingresos") if get(ra, "ingresos") else 0

    indicadores.append({
        "categoria":   "Rentabilidad",
        "indicador":   "Margen Neto",
        "formula":     "Utilidad Neta / Ingresos",
        "valor":       margen_neto,
        "valor_ant":   margen_neto_ant,
        "formato":     "porcentaje",
        "meta":        ">= 5%",
        "semaforo":    "verde" if margen_neto >= 0.05 else "amarillo" if margen_neto >= 0.02 else "rojo",
        "interpretacion": "% de ingresos que queda como utilidad",
    })

    return indicadores


# ── Excel ──────────────────────────────────────────────────────────

def formato_valor(val, fmt):
    if fmt == "ratio":   return f"{val:.2f}x"
    if fmt == "porcentaje": return f"{val:.1%}"
    if fmt == "dias":    return f"{val:.0f} días"
    return str(val)


def generar_excel(indicadores, archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicadores"

    anchos = [18, 30, 38, 14, 14, 12, 10, 32]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(anchos))}1")
    ws["A1"] = f"INDICADORES FINANCIEROS — {date.today().strftime('%d/%m/%Y')}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 28

    # Resumen semáforo
    verdes    = sum(1 for i in indicadores if i["semaforo"] == "verde")
    amarillos = sum(1 for i in indicadores if i["semaforo"] == "amarillo")
    rojos     = sum(1 for i in indicadores if i["semaforo"] == "rojo")

    ws.merge_cells(f"A2:{get_column_letter(len(anchos))}2")
    ws["A2"] = (f"✅ {verdes} indicadores OK   "
                f"⚠️  {amarillos} en atención   "
                f"🔴 {rojos} críticos")
    ws["A2"].font      = Font(name="Arial", bold=True, size=10)
    ws["A2"].alignment = CENTRO

    headers = ["Categoría", "Indicador", "Fórmula", "Valor Actual",
               "Valor Anterior", "Variación", "Meta", "Interpretación"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    cat_actual = None
    for i, ind in enumerate(indicadores, 4):
        semaforo = ind["semaforo"]
        color    = VERDE if semaforo == "verde" else AMARILLO if semaforo == "amarillo" else ROJO
        color_t  = VERDE_T if semaforo == "verde" else AMARILLO_T if semaforo == "amarillo" else ROJO_T

        # Variación
        if ind["valor_ant"] and ind["valor_ant"] != 0:
            variacion = (ind["valor"] - ind["valor_ant"]) / abs(ind["valor_ant"])
            var_txt   = f"{variacion:+.1%}"
        else:
            var_txt = "—"

        datos = [
            ind["categoria"] if ind["categoria"] != cat_actual else "",
            ind["indicador"],
            ind["formula"],
            formato_valor(ind["valor"], ind["formato"]),
            formato_valor(ind["valor_ant"], ind["formato"]),
            var_txt,
            ind["meta"],
            ind["interpretacion"],
        ]
        cat_actual = ind["categoria"]

        alns = [CENTRO, IZQUIERDA, IZQUIERDA, CENTRO, CENTRO, CENTRO, CENTRO, IZQUIERDA]
        for col, (v, a) in enumerate(zip(datos, alns), 1):
            c = ws.cell(i, col, v)
            c.font      = Font(name="Arial", bold=(col == 2), size=10, color=color_t)
            c.fill      = _fill(color)
            c.alignment = a
            c.border    = _borde()

    ws.freeze_panes = "A4"

    # ── Hoja 2: Semáforo visual ──
    ws2 = wb.create_sheet("Semáforo")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 14

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "SEMÁFORO DE INDICADORES"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws2["A1"].alignment = CENTRO
    ws2.row_dimensions[1].height = 28

    for col, h in enumerate(["Categoría", "Indicador", "Valor", "Estado", "Meta"], 1):
        c = ws2.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    iconos = {"verde": "✅", "amarillo": "⚠️", "rojo": "🔴"}
    for i, ind in enumerate(indicadores, 3):
        semaforo = ind["semaforo"]
        color    = VERDE if semaforo == "verde" else AMARILLO if semaforo == "amarillo" else ROJO
        datos    = [ind["categoria"], ind["indicador"],
                    formato_valor(ind["valor"], ind["formato"]),
                    iconos[semaforo], ind["meta"]]
        alns     = [CENTRO, IZQUIERDA, CENTRO, CENTRO, CENTRO]
        for col, (v, a) in enumerate(zip(datos, alns), 1):
            c = ws2.cell(i, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), _fill(color), a, _borde())

    wb.save(archivo)
    return archivo


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== INDICADORES FINANCIEROS ===\n")

    ruta = input("  Ruta archivo Excel (Balance + Resultados): ").strip()

    datos = cargar_datos(ruta)
    if datos[0] is None:
        return

    (balance, balance_ant), (resultados, resultados_ant) = datos

    indicadores = calcular_indicadores(balance, balance_ant, resultados, resultados_ant)

    print(f"\n  {'Indicador':<35} {'Valor':>12}  {'Semáforo'}")
    print(f"  {'-'*60}")
    for ind in indicadores:
        icono = "✅" if ind["semaforo"] == "verde" else "⚠️ " if ind["semaforo"] == "amarillo" else "🔴"
        print(f"  {ind['indicador']:<35} {formato_valor(ind['valor'], ind['formato']):>12}  {icono}")

    archivo = f"indicadores_financieros_{date.today().strftime('%Y%m%d')}.xlsx"
    salida  = generar_excel(indicadores, archivo)
    print(f"\n  ✅ Excel generado: {salida}\n")