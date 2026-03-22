"""
Gestión de Líneas de Crédito
Genera plantilla, procesa y reporta el estado de líneas de crédito bancarias.

Columnas requeridas en la plantilla:
- banco, tipo_linea, cupo_total, cupo_usado,
  tasa_anual, fecha_vencimiento, garantia
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center")
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")

TIPOS_LINEA = [
    "Sobregiro",
    "Crédito Rotativo",
    "Factoring",
    "Leasing",
    "Línea Capital de Trabajo",
    "Confirming",
]

COLORES_USO = {
    "bajo":   "E2EFDA",   # < 50%
    "medio":  "FFF2CC",   # 50% - 80%
    "alto":   "FCE4D6",   # 80% - 95%
    "critico": "FFCCCC",  # > 95%
}

COLORES_VENC = {
    "ok":      "E2EFDA",   # > 90 días
    "proximo": "FFF2CC",   # 30 - 90 días
    "urgente": "FCE4D6",   # 15 - 30 días
    "critico": "FFCCCC",   # < 15 días
}


# ── Plantilla ──────────────────────────────────────────────────────

def generar_plantilla(archivo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Líneas de Crédito"

    anchos = [22, 24, 16, 16, 10, 16, 24]
    headers = ["banco", "tipo_linea", "cupo_total", "cupo_usado",
               "tasa_anual", "fecha_vencimiento", "garantia"]

    for i, (h, w) in enumerate(zip(headers, anchos), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(1, i, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    # Filas de ejemplo
    ejemplos = [
        ("Banco BCI",       "Sobregiro",              50000000,  20000000, 8.5,  "31/12/2025", "Sin garantía"),
        ("Banco Santander", "Crédito Rotativo",       80000000,  65000000, 7.2,  "30/06/2025", "Pagaré"),
        ("Banco de Chile",  "Factoring",             120000000,  90000000, 9.0,  "31/03/2026", "Facturas cedidas"),
        ("Scotiabank",      "Línea Capital de Trabajo",60000000, 15000000, 6.8,  "31/08/2025", "Hipoteca"),
        ("Banco Estado",    "Leasing",                40000000,  40000000, 7.5,  "15/04/2025", "Bien arrendado"),
    ]

    for fila, datos in enumerate(ejemplos, 2):
        color = "F5F5F5" if fila % 2 == 0 else "FFFFFF"
        for col, val in enumerate(datos, 1):
            c = ws.cell(fila, col, val)
            c.font      = Font(name="Arial", size=10, color="0000FF")
            c.fill      = _fill(color)
            c.alignment = DERECHA if col in [3, 4, 5] else CENTRO if col == 6 else IZQUIERDA
            c.border    = _borde()
            if col in [3, 4]:
                c.number_format = "#,##0"
            if col == 5:
                c.number_format = "0.00"

    # Nota
    ws.cell(len(ejemplos) + 3, 1, "* tasa_anual en porcentaje (ej: 8.5 = 8.5%)").font = Font(
        name="Arial", size=9, color="888888", italic=True)
    ws.cell(len(ejemplos) + 4, 1, "* fecha_vencimiento en formato DD/MM/YYYY").font = Font(
        name="Arial", size=9, color="888888", italic=True)

    wb.save(archivo)
    return archivo


# ── Lógica ─────────────────────────────────────────────────────────

def cargar_archivo(ruta):
    try:
        df = pd.read_excel(ruta)
    except FileNotFoundError:
        print(f"\n  ❌ Archivo no encontrado: {ruta}\n")
        return None

    df.columns = df.columns.str.strip().str.lower()

    for col in ["banco", "tipo_linea", "cupo_total", "cupo_usado", "fecha_vencimiento"]:
        if col not in df.columns:
            print(f"\n  ❌ Falta la columna '{col}'.\n")
            return None

    df["fecha_vencimiento"] = pd.to_datetime(df["fecha_vencimiento"], dayfirst=True)
    df["cupo_total"]        = pd.to_numeric(df["cupo_total"],  errors="coerce")
    df["cupo_usado"]        = pd.to_numeric(df["cupo_usado"],  errors="coerce")
    df["tasa_anual"]        = pd.to_numeric(df.get("tasa_anual", 0), errors="coerce").fillna(0)
    df["garantia"]          = df.get("garantia", "—").fillna("—")

    hoy = pd.Timestamp(date.today())
    df["cupo_disponible"]   = df["cupo_total"] - df["cupo_usado"]
    df["pct_uso"]           = df["cupo_usado"] / df["cupo_total"]
    df["dias_vencimiento"]  = (df["fecha_vencimiento"] - hoy).dt.days
    df["costo_mensual"]     = df["cupo_usado"] * (df["tasa_anual"] / 100 / 12)

    def semaforo_uso(pct):
        if pct < 0.50:   return "bajo"
        elif pct < 0.80: return "medio"
        elif pct < 0.95: return "alto"
        else:            return "critico"

    def semaforo_venc(dias):
        if dias > 90:    return "ok"
        elif dias > 30:  return "proximo"
        elif dias > 15:  return "urgente"
        else:            return "critico"

    df["sem_uso"]  = df["pct_uso"].apply(semaforo_uso)
    df["sem_venc"] = df["dias_vencimiento"].apply(semaforo_venc)

    return df


def concentracion_banco(df):
    total = df["cupo_total"].sum()
    return df.groupby("banco").agg(
        lineas          = ("tipo_linea",      "count"),
        cupo_total      = ("cupo_total",      "sum"),
        cupo_usado      = ("cupo_usado",      "sum"),
        cupo_disponible = ("cupo_disponible", "sum"),
        costo_mensual   = ("costo_mensual",   "sum"),
    ).assign(
        concentracion = lambda x: x["cupo_total"] / total,
        pct_uso       = lambda x: x["cupo_usado"] / x["cupo_total"],
    ).sort_values("cupo_total", ascending=False).reset_index()


# ── Excel ──────────────────────────────────────────────────────────

def generar_reporte(df, df_bancos, archivo):
    wb = Workbook()
    hoy = date.today().strftime("%d/%m/%Y")

    total_cupo  = df["cupo_total"].sum()
    total_usado = df["cupo_usado"].sum()
    total_disp  = df["cupo_disponible"].sum()
    total_costo = df["costo_mensual"].sum()
    pct_uso_global = total_usado / total_cupo if total_cupo else 0

    # ── Hoja 1: Resumen ejecutivo ──
    ws = wb.active
    ws.title = "Resumen"

    for i, w in enumerate([22, 24, 16, 16, 10, 16, 14, 14, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:J1")
    ws["A1"] = f"GESTIÓN DE LÍNEAS DE CRÉDITO — {hoy}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = (f"Cupo total: ${total_cupo:,.0f}   |   "
                f"Usado: ${total_usado:,.0f} ({pct_uso_global:.1%})   |   "
                f"Disponible: ${total_disp:,.0f}   |   "
                f"Costo mensual: ${total_costo:,.0f}")
    ws["A2"].font      = Font(name="Arial", size=9, color="555555")
    ws["A2"].alignment = CENTRO

    headers = ["Banco", "Tipo Línea", "Cupo Total", "Cupo Usado",
               "Tasa %", "Disponible", "% Uso", "Vence",
               "Días", "Garantía"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(3, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    df_sorted = df.sort_values("pct_uso", ascending=False)
    for i, (_, row) in enumerate(df_sorted.iterrows(), 4):
        color_uso  = COLORES_USO.get(row["sem_uso"],  "FFFFFF")
        color_venc = COLORES_VENC.get(row["sem_venc"], "FFFFFF")

        datos = [
            row["banco"], row["tipo_linea"],
            row["cupo_total"], row["cupo_usado"],
            row["tasa_anual"], row["cupo_disponible"],
            row["pct_uso"],
            row["fecha_vencimiento"].strftime("%d/%m/%Y") if pd.notna(row["fecha_vencimiento"]) else "—",
            int(row["dias_vencimiento"]) if pd.notna(row["dias_vencimiento"]) else 0,
            row["garantia"],
        ]
        fmts = [None, None, "#,##0", "#,##0", "0.00", "#,##0", "0.0%", None, "#,##0", None]
        alns = [IZQUIERDA, IZQUIERDA, DERECHA, DERECHA, CENTRO,
                DERECHA, CENTRO, CENTRO, CENTRO, IZQUIERDA]

        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            # Colorear según uso o vencimiento
            if col in [7]:
                color = color_uso
            elif col in [8, 9]:
                color = color_venc
            else:
                color = color_uso

            c = ws.cell(i, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), _fill(color), a, _borde())
            if f:
                c.number_format = f

    # Totales
    tr = len(df) + 4
    ws.cell(tr, 1, "TOTAL").font      = Font(name="Arial", bold=True, size=10)
    ws.cell(tr, 1).fill               = _fill("D9D9D9")
    ws.cell(tr, 1).alignment          = CENTRO
    ws.cell(tr, 1).border             = _borde()
    for col in range(2, 11):
        c = ws.cell(tr, col)
        c.fill, c.border = _fill("D9D9D9"), _borde()
        c.font = Font(name="Arial", bold=True, size=10)
        if col in [3, 4, 6]:
            letra           = get_column_letter(col)
            c.value         = f"=SUM({letra}4:{letra}{tr-1})"
            c.number_format = "#,##0"
            c.alignment     = DERECHA
        if col == 7:
            c.value         = f"={get_column_letter(4)}{tr}/{get_column_letter(3)}{tr}"
            c.number_format = "0.0%"
            c.alignment     = CENTRO

    ws.freeze_panes = "A4"

    # ── Hoja 2: Concentración por banco ──
    ws2 = wb.create_sheet("Por Banco")
    for i, w in enumerate([22, 8, 16, 16, 16, 16, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "CONCENTRACIÓN POR BANCO"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws2["A1"].alignment = CENTRO
    ws2.row_dimensions[1].height = 28

    headers2 = ["Banco", "Líneas", "Cupo Total", "Cupo Usado",
                 "Disponible", "Costo Mensual", "Concentración"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    for i, (_, row) in enumerate(df_bancos.iterrows(), 3):
        conc    = row["concentracion"]
        color   = "FCE4D6" if conc >= 0.40 else "FFF2CC" if conc >= 0.25 else "E2EFDA"
        datos   = [row["banco"], int(row["lineas"]), row["cupo_total"],
                   row["cupo_usado"], row["cupo_disponible"],
                   row["costo_mensual"], row["concentracion"]]
        fmts    = [None, "#,##0", "#,##0", "#,##0", "#,##0", "#,##0", "0.0%"]
        alns    = [IZQUIERDA, CENTRO, DERECHA, DERECHA, DERECHA, DERECHA, CENTRO]
        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws2.cell(i, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), _fill(color), a, _borde())
            if f:
                c.number_format = f

    # ── Hoja 3: Alertas ──
    ws3 = wb.create_sheet("Alertas")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 10
    ws3.column_dimensions["E"].width = 30

    ws3.merge_cells("A1:E1")
    ws3["A1"] = f"ALERTAS — {hoy}"
    ws3["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws3["A1"].alignment = CENTRO
    ws3.row_dimensions[1].height = 28

    for col, h in enumerate(["Banco", "Tipo Línea", "Valor", "Prioridad", "Alerta"], 1):
        c = ws3.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    alertas = []
    for _, row in df.iterrows():
        if row["sem_uso"] in ["alto", "critico"]:
            alertas.append({
                "banco":     row["banco"],
                "linea":     row["tipo_linea"],
                "valor":     f"{row['pct_uso']:.1%} usado",
                "prioridad": "🔴 Alta" if row["sem_uso"] == "critico" else "⚠️  Media",
                "mensaje":   f"Cupo casi agotado — disponible solo ${row['cupo_disponible']:,.0f}",
                "color":     "FFCCCC" if row["sem_uso"] == "critico" else "FCE4D6",
            })
        if row["sem_venc"] in ["urgente", "critico"]:
            alertas.append({
                "banco":     row["banco"],
                "linea":     row["tipo_linea"],
                "valor":     f"{int(row['dias_vencimiento']) if pd.notna(row['dias_vencimiento']) else 0} días",
                "prioridad": "🔴 Alta" if row["sem_venc"] == "critico" else "⚠️  Media",
                "mensaje":   f"Línea vence el {row['fecha_vencimiento'].strftime('%d/%m/%Y') if pd.notna(row['fecha_vencimiento']) else '—'} — gestionar renovación",
                "color":     "FFCCCC" if row["sem_venc"] == "critico" else "FCE4D6",
            })

    if not alertas:
        ws3.merge_cells("A3:E3")
        ws3["A3"] = "✅ Sin alertas activas"
        ws3["A3"].font      = Font(name="Arial", size=10, color="375623")
        ws3["A3"].alignment = CENTRO
    else:
        for i, alerta in enumerate(alertas, 3):
            datos = [alerta["banco"], alerta["linea"], alerta["valor"],
                     alerta["prioridad"], alerta["mensaje"]]
            alns  = [IZQUIERDA, IZQUIERDA, CENTRO, CENTRO, IZQUIERDA]
            for col, (v, a) in enumerate(zip(datos, alns), 1):
                c = ws3.cell(i, col, v)
                c.font, c.fill, c.alignment, c.border = (
                    Font(name="Arial", size=10), _fill(alerta["color"]), a, _borde())

    wb.save(archivo)
    return archivo


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== GESTIÓN DE LÍNEAS DE CRÉDITO ===\n")
    print("  [1] Generar plantilla")
    print("  [2] Procesar y generar reporte\n")

    opcion = input("  Selecciona una opción [1-2]: ").strip()

    if opcion == "1":
        archivo = "data/lineas_credito.xlsx"
        generar_plantilla(archivo)
        print(f"\n  ✅ Plantilla generada: {archivo}")
        print(f"  👉 Completa los datos y vuelve con opción 2.\n")

    elif opcion == "2":
        ruta = input("  Ruta archivo de líneas (.xlsx): ").strip()
        df   = cargar_archivo(ruta)
        if df is None:
            return

        print(f"\n  Líneas cargadas      : {len(df)}")
        print(f"  Bancos               : {df['banco'].nunique()}")
        print(f"  Cupo total           : ${df['cupo_total'].sum():,.0f}")
        print(f"  Cupo usado           : ${df['cupo_usado'].sum():,.0f} ({df['cupo_usado'].sum()/df['cupo_total'].sum():.1%})")
        print(f"  Cupo disponible      : ${df['cupo_disponible'].sum():,.0f}")
        print(f"  Costo mensual total  : ${df['costo_mensual'].sum():,.0f}\n")

        alertas_uso  = df[df["sem_uso"].isin(["alto", "critico"])]
        alertas_venc = df[df["sem_venc"].isin(["urgente", "critico"])]

        if len(alertas_uso) > 0:
            print(f"  ⚠️  {len(alertas_uso)} línea(s) con cupo casi agotado")
        if len(alertas_venc) > 0:
            print(f"  ⚠️  {len(alertas_venc)} línea(s) con vencimiento próximo")

        df_bancos = concentracion_banco(df)
        archivo   = f"reporte_lineas_credito_{date.today().strftime('%Y%m%d')}.xlsx"
        salida    = generar_reporte(df, df_bancos, archivo)
        print(f"\n  ✅ Excel generado: {salida}\n")

    else:
        print("\n  Opción no válida.\n")