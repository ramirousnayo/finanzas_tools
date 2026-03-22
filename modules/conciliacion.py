"""
Conciliación Bancaria
Cruza extracto del banco vs libro interno por fecha y monto.

Columnas requeridas:
- Extracto banco  : fecha, monto, descripcion
- Libro interno   : fecha, monto, descripcion
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center")
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")


# ── Lógica ─────────────────────────────────────────────────────────

def cargar_archivo(ruta, nombre):
    try:
        df = pd.read_excel(ruta)
    except FileNotFoundError:
        print(f"\n  ❌ Archivo no encontrado: {ruta}\n")
        return None

    df.columns = df.columns.str.strip().str.lower()

    columnas_req = ["fecha", "monto", "descripcion"]
    for col in columnas_req:
        if col not in df.columns:
            print(f"\n  ❌ Falta la columna '{col}' en {nombre}.")
            print(f"  Columnas encontradas: {list(df.columns)}\n")
            return None

    df["fecha"] = pd.to_datetime(df["fecha"], dayfirst=True)
    df["monto"] = pd.to_numeric(df["monto"], errors="coerce")
    df = df.dropna(subset=["fecha", "monto"])
    return df


def conciliar(df_banco, df_interno):
    # Crear clave de conciliación: fecha + monto
    df_banco["clave"]    = df_banco["fecha"].dt.strftime("%Y-%m-%d") + "_" + df_banco["monto"].astype(str)
    df_interno["clave"]  = df_interno["fecha"].dt.strftime("%Y-%m-%d") + "_" + df_interno["monto"].astype(str)

    claves_banco    = set(df_banco["clave"])
    claves_interno  = set(df_interno["clave"])

    # Matches — aparecen en ambos
    claves_match    = claves_banco & claves_interno

    # Solo en banco — no están en el libro interno
    solo_banco      = claves_banco - claves_interno

    # Solo en interno — no están en el extracto banco
    solo_interno    = claves_interno - claves_banco

    df_match        = df_banco[df_banco["clave"].isin(claves_match)].copy()
    df_solo_banco   = df_banco[df_banco["clave"].isin(solo_banco)].copy()
    df_solo_interno = df_interno[df_interno["clave"].isin(solo_interno)].copy()

    return df_match, df_solo_banco, df_solo_interno


# ── Excel ──────────────────────────────────────────────────────────

def escribir_hoja(wb, titulo, df, color_fondo, color_header, nombre_hoja):
    ws = wb.create_sheet(nombre_hoja)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30

    ws.merge_cells("A1:C1")
    ws["A1"] = titulo
    ws["A1"].font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ws["A1"].fill      = _fill(color_header)
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 24

    headers = ["Fecha", "Monto", "Descripción"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill(color_header)
        c.alignment = CENTRO
        c.border    = _borde()

    if df.empty:
        ws.merge_cells("A3:C3")
        ws["A3"] = "Sin registros"
        ws["A3"].font      = Font(name="Arial", size=10, color="888888")
        ws["A3"].alignment = CENTRO
        return

    for i, (_, row) in enumerate(df.iterrows(), 1):
        fila    = i + 2
        relleno = _fill(color_fondo) if i % 2 == 0 else _fill("FFFFFF")
        datos   = [row["fecha"].strftime("%d/%m/%Y"), row["monto"], row.get("descripcion", "")]
        fmts    = ["DD/MM/YYYY", "#,##0", None]
        alns    = [CENTRO, DERECHA, IZQUIERDA]
        for col, (v, f, a) in enumerate(zip(datos, fmts, alns), 1):
            c = ws.cell(fila, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), relleno, a, _borde())
            if f:
                c.number_format = f

    # Total
    tr = len(df) + 3
    ws[f"A{tr}"] = "TOTAL"
    ws[f"A{tr}"].font      = Font(name="Arial", bold=True, size=10)
    ws[f"A{tr}"].fill      = _fill("D9D9D9")
    ws[f"A{tr}"].alignment = CENTRO
    ws[f"A{tr}"].border    = _borde()
    c = ws.cell(tr, 2)
    c.value         = f"=SUM(B3:B{tr-1})"
    c.font          = Font(name="Arial", bold=True, size=10)
    c.fill          = _fill("D9D9D9")
    c.alignment     = DERECHA
    c.border        = _borde()
    c.number_format = "#,##0"
    ws.cell(tr, 3).fill   = _fill("D9D9D9")
    ws.cell(tr, 3).border = _borde()

    ws.freeze_panes = "A3"


def generar_excel(df_match, df_solo_banco, df_solo_interno, archivo):
    wb = Workbook()

    # Hoja resumen
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18

    ws.merge_cells("A1:C1")
    ws["A1"] = f"RESUMEN CONCILIACIÓN BANCARIA — {date.today().strftime('%d/%m/%Y')}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 28

    headers = ["Concepto", "Registros", "Monto Total"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    resumen = [
        ("✅ Conciliados",              len(df_match),        df_match["monto"].sum()        if not df_match.empty else 0,        "C6EFCE", "375623"),
        ("⚠️  Solo en banco",           len(df_solo_banco),   df_solo_banco["monto"].sum()   if not df_solo_banco.empty else 0,   "FFF2CC", "7F6000"),
        ("🔴 Solo en libro interno",    len(df_solo_interno), df_solo_interno["monto"].sum() if not df_solo_interno.empty else 0, "FFCCCC", "C00000"),
    ]

    for i, (concepto, cant, monto, color, color_txt) in enumerate(resumen, 3):
        ws.cell(i, 1, concepto).font      = Font(name="Arial", bold=True, size=10, color=color_txt)
        ws.cell(i, 1).fill      = _fill(color)
        ws.cell(i, 1).alignment = IZQUIERDA
        ws.cell(i, 1).border    = _borde()
        ws.cell(i, 2, cant).font          = Font(name="Arial", bold=True, size=10, color=color_txt)
        ws.cell(i, 2).fill      = _fill(color)
        ws.cell(i, 2).alignment = CENTRO
        ws.cell(i, 2).border    = _borde()
        c = ws.cell(i, 3, monto)
        c.font          = Font(name="Arial", bold=True, size=10, color=color_txt)
        c.fill          = _fill(color)
        c.alignment     = DERECHA
        c.border        = _borde()
        c.number_format = "#,##0"

    # Hojas de detalle
    escribir_hoja(wb, "✅ Movimientos Conciliados",       df_match,        "D6E4F0", "375623", "Conciliados")
    escribir_hoja(wb, "⚠️  Solo en Banco",                df_solo_banco,   "FFF2CC", "7F6000", "Solo en Banco")
    escribir_hoja(wb, "🔴 Solo en Libro Interno",         df_solo_interno, "FFCCCC", "C00000", "Solo en Interno")

    wb.save(archivo)
    return archivo


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== CONCILIACIÓN BANCARIA ===\n")

    ruta_banco   = input("  Ruta extracto banco    : ").strip()
    ruta_interno = input("  Ruta libro interno     : ").strip()

    df_banco   = cargar_archivo(ruta_banco,   "extracto banco")
    df_interno = cargar_archivo(ruta_interno, "libro interno")

    if df_banco is None or df_interno is None:
        return

    print(f"\n  Registros banco    : {len(df_banco)}")
    print(f"  Registros internos : {len(df_interno)}")

    df_match, df_solo_banco, df_solo_interno = conciliar(df_banco, df_interno)

    total = len(df_banco)
    pct   = len(df_match) / total * 100 if total > 0 else 0

    print(f"\n  ✅ Conciliados       : {len(df_match)} ({pct:.1f}%)")
    print(f"  ⚠️  Solo en banco    : {len(df_solo_banco)}")
    print(f"  🔴 Solo en interno   : {len(df_solo_interno)}\n")

    if len(df_solo_banco) > 0 or len(df_solo_interno) > 0:
        print(f"  ⚠️  Hay {len(df_solo_banco) + len(df_solo_interno)} movimientos sin conciliar — revisa el detalle en el Excel.\n")

    archivo = f"conciliacion_{date.today().strftime('%Y%m%d')}.xlsx"
    salida  = generar_excel(df_match, df_solo_banco, df_solo_interno, archivo)
    print(f"  ✅ Excel generado: {salida}\n")