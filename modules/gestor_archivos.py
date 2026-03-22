"""
Gestor de Archivos Financieros
Organiza, renombra y genera índice Excel de archivos del área de finanzas.

Estructura de carpetas generada:
    destino/
    └── 2025/
        └── 03_Marzo/
            ├── Contratos/
            ├── Bancos/
            ├── Respaldos/
            ├── Prestamos/
            └── Facturas/
"""

import os
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Estilos ────────────────────────────────────────────────────────

def _fill(hex):
    return PatternFill("solid", fgColor=hex)

def _borde():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTRO    = Alignment(horizontal="center", vertical="center")
DERECHA   = Alignment(horizontal="right",  vertical="center")
IZQUIERDA = Alignment(horizontal="left",   vertical="center")

MESES = {
    1: "01_Enero",   2: "02_Febrero", 3: "03_Marzo",
    4: "04_Abril",   5: "05_Mayo",    6: "06_Junio",
    7: "07_Julio",   8: "08_Agosto",  9: "09_Septiembre",
    10: "10_Octubre",11: "11_Noviembre",12: "12_Diciembre"
}

CATEGORIAS = {
    "contrato":  "Contratos",
    "banco":     "Bancos",
    "respaldo":  "Respaldos",
    "prestamo":  "Prestamos",
    "factura":   "Facturas",
}

COLORES_CAT = {
    "Contratos": "D6E4F0",
    "Bancos":    "E2EFDA",
    "Respaldos": "FFF2CC",
    "Prestamos": "FCE4D6",
    "Facturas":  "F2F2F2",
}


# ── Lógica ─────────────────────────────────────────────────────────

def detectar_categoria(nombre_archivo):
    """Detecta la categoría según palabras clave en el nombre del archivo."""
    nombre = nombre_archivo.lower()
    for clave, categoria in CATEGORIAS.items():
        if clave in nombre:
            return categoria
    return "Respaldos"  # categoría por defecto


def obtener_fecha_archivo(ruta):
    """Obtiene la fecha de modificación del archivo."""
    timestamp = os.path.getmtime(ruta)
    return datetime.fromtimestamp(timestamp)


def construir_nombre(categoria, fecha, nombre_original):
    """Renombra el archivo con convención: CATEGORIA_YYYYMM_nombre_original."""
    ext       = Path(nombre_original).suffix
    base      = Path(nombre_original).stem
    fecha_str = fecha.strftime("%Y%m")
    return f"{categoria.upper()}_{fecha_str}_{base}{ext}"


def crear_estructura(destino, anio, mes):
    """Crea la estructura de carpetas si no existe."""
    mes_nombre = MESES[mes]
    for categoria in CATEGORIAS.values():
        carpeta = Path(destino) / str(anio) / mes_nombre / categoria
        carpeta.mkdir(parents=True, exist_ok=True)
    return Path(destino) / str(anio) / mes_nombre


def organizar_archivos(origen, destino, anio, mes):
    """Organiza y renombra los archivos del directorio origen."""
    origen_path = Path(origen)
    registros   = []

    if not origen_path.exists():
        print(f"\n  ❌ Carpeta no encontrada: {origen}\n")
        return None

    archivos = [f for f in origen_path.iterdir()
                if f.is_file() and not f.name.startswith(".")]

    if not archivos:
        print(f"\n  ❌ No hay archivos en: {origen}\n")
        return None

    base_destino = crear_estructura(destino, anio, mes)

    for archivo in archivos:
        categoria     = detectar_categoria(archivo.name)
        fecha         = obtener_fecha_archivo(archivo)
        nombre_nuevo  = construir_nombre(categoria, fecha, archivo.name)
        carpeta_dest  = base_destino / categoria
        ruta_destino  = carpeta_dest / nombre_nuevo

        # Si ya existe, agregar sufijo
        contador = 1
        while ruta_destino.exists():
            stem     = Path(nombre_nuevo).stem
            ext      = Path(nombre_nuevo).suffix
            ruta_destino = carpeta_dest / f"{stem}_{contador}{ext}"
            contador += 1

        shutil.copy2(archivo, ruta_destino)

        registros.append({
            "nombre_original": archivo.name,
            "nombre_nuevo":    ruta_destino.name,
            "categoria":       categoria,
            "fecha":           fecha.strftime("%d/%m/%Y"),
            "tamaño":          f"{archivo.stat().st_size / 1024:.1f} KB",
            "ruta":            str(ruta_destino),
        })

        print(f"  → {archivo.name[:35]:<35} → {categoria}/{ruta_destino.name}")

    return registros


# ── Excel ──────────────────────────────────────────────────────────

def generar_indice(registros, archivo_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Índice"

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 50

    ws.merge_cells("A1:F1")
    ws["A1"] = f"ÍNDICE DE ARCHIVOS — Generado {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws["A1"].alignment = CENTRO
    ws.row_dimensions[1].height = 28

    headers = ["Nombre Original", "Nombre Nuevo", "Categoría", "Fecha Archivo", "Tamaño", "Ruta"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    for i, reg in enumerate(registros, 1):
        fila    = i + 2
        color   = COLORES_CAT.get(reg["categoria"], "FFFFFF")
        relleno = _fill(color)
        datos   = [reg["nombre_original"], reg["nombre_nuevo"],
                   reg["categoria"], reg["fecha"],
                   reg["tamaño"], reg["ruta"]]
        alns    = [IZQUIERDA, IZQUIERDA, CENTRO, CENTRO, CENTRO, IZQUIERDA]

        for col, (v, a) in enumerate(zip(datos, alns), 1):
            c = ws.cell(fila, col, v)
            c.font, c.fill, c.alignment, c.border = (
                Font(name="Arial", size=10), relleno, a, _borde())

    # Resumen por categoría
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "RESUMEN POR CATEGORÍA"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=13, color="1F3864")
    ws2["A1"].alignment = CENTRO

    for col, h in enumerate(["Categoría", "Archivos", "% del Total"], 1):
        c = ws2.cell(2, col, h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = _fill("1F3864")
        c.alignment = CENTRO
        c.border    = _borde()

    total = len(registros)
    conteo = {}
    for reg in registros:
        conteo[reg["categoria"]] = conteo.get(reg["categoria"], 0) + 1

    for i, (cat, cant) in enumerate(sorted(conteo.items()), 3):
        pct     = cant / total if total > 0 else 0
        color   = COLORES_CAT.get(cat, "FFFFFF")
        ws2.cell(i, 1, cat).font       = Font(name="Arial", bold=True, size=10)
        ws2.cell(i, 1).fill            = _fill(color)
        ws2.cell(i, 1).alignment       = CENTRO
        ws2.cell(i, 1).border          = _borde()
        ws2.cell(i, 2, cant).font      = Font(name="Arial", size=10)
        ws2.cell(i, 2).fill            = _fill(color)
        ws2.cell(i, 2).alignment       = CENTRO
        ws2.cell(i, 2).border          = _borde()
        c = ws2.cell(i, 3, pct)
        c.font, c.fill, c.alignment, c.border = (
            Font(name="Arial", size=10), _fill(color), CENTRO, _borde())
        c.number_format = "0.0%"

    wb.save(archivo_salida)
    return archivo_salida


# ── Menú ───────────────────────────────────────────────────────────

def run():
    print("\n=== GESTOR DE ARCHIVOS FINANCIEROS ===\n")

    origen  = input("  Carpeta origen (archivos a organizar) : ").strip()
    destino = input("  Carpeta destino                       : ").strip() or "archivos_organizados"
    anio    = int(input("  Año  (ej: 2025)                       : ") or datetime.now().year)
    mes     = int(input("  Mes  (1-12)                           : ") or datetime.now().month)

    if mes < 1 or mes > 12:
        print("\n  ❌ Mes inválido.\n")
        return

    print(f"\n  Organizando archivos...\n")
    registros = organizar_archivos(origen, destino, anio, mes)

    if not registros:
        return

    print(f"\n  Total archivos procesados : {len(registros)}")
    for cat, cant in sorted({r["categoria"]: 0 for r in registros}.items()):
        cant = sum(1 for r in registros if r["categoria"] == cat)
        print(f"  {cat:<14} : {cant} archivos")

    archivo = f"indice_archivos_{anio}{mes:02d}.xlsx"
    generar_indice(registros, archivo)
    print(f"\n  ✅ Índice generado: {archivo}\n")